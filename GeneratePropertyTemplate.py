import requests
import csv
from common import *
from config import config ,load_employee_creation_config
from common import superuser_login
from config import config
import io
import os
import numpy
import pandas as pd
import openpyxl
from openpyxl import Workbook, utils
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.comments import Comment
from datetime import datetime, timedelta
from math import isnan

# import shutil
# import xlrd
# import requests
# import xlwt
# from xlwt import Workbook
# import xlsxwriter

def main():
    Flag =False
    tenantMapping={}
    post_data_resp_list=[]

    templateFile = r"D:\eGov\Data\WS\Template\Template for Existing Property Detail.xlsx"
    df1 = pd.read_excel(templateFile, 'Property Ownership Details')
    df2 = pd.read_excel(templateFile, 'Master Data')
    df3 = pd.read_excel(templateFile, 'Master_UsageType')
    
    count = 0
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            tenantMapping[module["code"].lower()]=module["code"].lower()[3:]

    for root, dirs, files in os.walk(r"D:\eGov\Data\WS\ABASPY", topdown=True):
        for name in dirs:
            # print(dirs)
            # print (os.path.join(root, name))
            # print(root)
            # print(name)            
            subfolder = os.path.join(root, name)
            # print(subfolder)            
            cbFile =os.path.join(root, name,"BEL_Template for Existing Property Detail.xlsx")
            # print(cbFile)
            
            city = subfolder.replace(r"D:\eGov\Data\WS\ABASPY\CB ","" ).strip().lower()
            city = "pb." + city

            if city not in tenantMapping:
                print("Not In city",city)
                continue
            cityname = tenantMapping[city]

            if os.path.exists(cbFile) :  
                template_path = os.path.join(r"D:/eGov/Data/WS/Template/Property1/CB " + cityname) 
                # template_file = os.path.join(config.LOG_PATH ,  "Locality.xlsx" )
                dfLocality = getLocalityData(cityname)
                workbook1 = openpyxl.load_workbook(cbFile)   
                writer = pd.ExcelWriter(cbFile, engine='openpyxl')   
                writer.book = workbook1                  
                sheet = workbook1.get_sheet_by_name('Property Assembly Detail')                
                if(ValidateCols(sheet) == False):
                    print(cityname, "Column Order Validation Failed")
                    continue
                df1.to_excel(writer,sheet_name="Property Ownership Details",index=False) 
                df2.to_excel(writer,sheet_name="Master Data",index=False) 
                df3.to_excel(writer,sheet_name="Master_UsageType",index=False) 
                dfLocality.to_excel(writer,sheet_name="Locality",index=False)            
                os.makedirs(template_path, exist_ok=True)
                insert_columns(sheet)  
                                             
                generatedFile = os.path.join(template_path,'Template for Existing Property-Integrated with ABAS-' + cityname + '.xlsx')
                workbook1.save(generatedFile)        
                workbook1.close()
                # add_header(templateFile, generatedFile) 
                DataValidation1(generatedFile)

                

                print(cityname, " Done")
                count = count + 1
            else:
                print(cityname, " file does not exist")            
                
    print("Total Count: ", count)  
            
def add_header(sheet1, sheet2):
    wb = openpyxl.load_workbook(r"D:\eGov\Data\WS\Template\Template for Existing Property Detail.xlsx")                               
    template_sheet1 = wb.get_sheet_by_name('Property Assembly Detail')
    template_sheet2 = wb.get_sheet_by_name('Property Ownership Details')  
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    column_list_sheet1 = [c.value for c in next(template_sheet1.iter_rows(min_row=1, max_row=1))]
    for col_num, value in enumerate(column_list_sheet1):
        sheet1.cell(row=1, column=col_num+1).value = value
        sheet1.cell(row=1, column=col_num+1).fill = PatternFill("solid", fgColor="07AEF9")
        sheet1.cell(row=1, column=col_num+1).font = Font(bold=True)
        sheet1.cell(row=1, column=col_num+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet1.cell(row=1, column=col_num+1).alignment = Alignment(wrap_text=True, horizontal="center")
    column_list_sheet1 = [c.value for c in next(template_sheet1.iter_rows(min_row=2, max_row=2))]
    for col_num, value in enumerate(column_list_sheet1):
        sheet1.cell(row=2, column=col_num+1).value = value
        sheet1.cell(row=2, column=col_num+1).fill = PatternFill("solid", fgColor="07AEF9")
        sheet1.cell(row=2, column=col_num+1).font = Font(bold=True)
        sheet1.cell(row=2, column=col_num+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet1.cell(row=2, column=col_num+1).alignment = Alignment(wrap_text=True, horizontal="center")

    column_list_sheet2 = [c.value for c in next(template_sheet2.iter_rows(min_row=1, max_row=1))]
    for col_num, value in enumerate(column_list_sheet2):
        sheet2.cell(row=1, column=col_num+1).value = value
        sheet2.cell(row=1, column=col_num+1).fill = PatternFill("solid", fgColor="07AEF9")
        sheet2.cell(row=1, column=col_num+1).font = Font(bold=True)
        sheet2.cell(row=1, column=col_num+1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        sheet2.cell(row=1, column=col_num+1).alignment = Alignment(wrap_text=True, horizontal="center")

    sheet1['AR3'].value = sheet2['N2'].value
    add_Comment(sheet1, sheet2)

def add_Comment(sheet1, sheet2):
    sheet1['F2'].comment = Comment('Applicable only in case of Property Type is Flat/Independent Building','')
    sheet1['I2'].comment = Comment('This field should be in reference with usage type. Suppose usage type is sleected as  "Commercial" then sub usage type should be any one of commercial category.','')
    sheet1['K2'].comment = Comment('Applicable only in case of Property type - Flat/Part of Building. Default Value - 1','')
    sheet1['AJ2'].comment = Comment('Does owner have the same sorrespondance address where the property located. If yes, then no need to fill correspondence address.','')
    sheet2['J2'].comment = Comment('Does owner have the same sorrespondance address where the property located. If yes, then no need to fill correspondence address.','')

def insert_columns(sheet):
    sheet.insert_rows(1)
    sheet.insert_cols(3)
    sheet.move_range("R1:R50000", rows=0, cols=-15, translate=True)   
    sheet.insert_cols(6, 3)
    sheet.move_range("R1:S50000", rows=0, cols=-11, translate=True)
    sheet.delete_cols(12)
    sheet.delete_cols(18, 5)
    sheet.insert_cols(4, 3) 
    sheet.insert_cols(8, 5)    
    sheet.insert_cols(18)   
    sheet.insert_cols(21)
    sheet.insert_cols(26, 3)
    sheet.insert_cols(30, 15)
    
    return

def ValidateCols(sheet):
    proper_column_order = ['Sl No.', 'Existing Property ID* ( Unique Value on which property are getting searched in existing system ) ',
     'Usage type *', 'CB Name *', 'Street Name*', 'House / Door No*', 'Pin Code*', 'Location', 'ARV', 'RV', 'Financial Year', 
     'Is Property on Dispute(Yes/No) ', 'Name*', 'Ward No', 'Block No', 'Location', 'Old PropertyCode']
    # print(proper_column_order)
    column_list = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    # print(column_list)
    validated = True
    for i in range(0, 16):
        if(proper_column_order[i].strip() != column_list[i].strip()) :
            print(column_list[i])
            validated = False
            break

    return validated    


def DataValidation1(generatedFile):
    # generatedFile = os.path.join(r"D:/eGov/Data/WS/Template/Property/CB " + 'agra' ,'Template for Existing Property-Integrated with ABAS-' + 'agra' + '.xlsx')
    workbook2 = openpyxl.load_workbook(generatedFile)   
    sheet = workbook2.get_sheet_by_name('Property Assembly Detail')    

    addValidationToColumns(sheet,"D","$A$2:$A$4","Master Data")
    addValidationToColumns(sheet,"H","$J$2:$J$9","Master_UsageType")
    # for I
    addValidationToColumns(sheet,"J","$O$2:$O$4","Master Data")
    addValidationToColumns(sheet,"N","$A$2:$A$500","Locality")
    addValidationToColumns(sheet,"U","$F$2:$F$4","Master Data")
    addValidationToColumns(sheet,"Y","$P$2:$P$3","Master Data")
    addValidationToColumns(sheet,"Z","$P$2:$P$3","Master Data")
    addValidationToColumns(sheet,"AA","$P$2:$P$3","Master Data")
    addValidationToColumns(sheet,"AB","$C$2:$C$5","Master Data")
    addValidationToColumns(sheet,"AF","$K$2:$K$4","Master Data")
    addValidationToColumns(sheet,"AI","$J$2:$J$4","Master Data")
    addValidationToColumns(sheet,"AJ","$P$2:$P$3","Master Data")
    addValidationToColumns(sheet,"AL","$H$2:$H$8","Master Data")
    addValidationToColumns(sheet,"AN","$D$2:$D$10","Master Data")

    sheet2 = workbook2.get_sheet_by_name('Property Ownership Details')
    addValidationToColumns1(sheet2,"B","$C$2:$C$5","Master Data")
    addValidationToColumns1(sheet2,"F","$K$2:$K$4","Master Data")
    addValidationToColumns1(sheet2,"I","$J$2:$J$4","Master Data")
    addValidationToColumns1(sheet2,"J","$P$2:$P$3","Master Data")
    addValidationToColumns1(sheet2,"L","$H$2:$H$8","Master Data")

    add_header(sheet, sheet2)
    sheet.merge_cells('B1:L1')
    sheet.merge_cells('M1:U1')
    sheet.merge_cells('V1:AA1')
    sheet.merge_cells('AB1:AP1')

    SubUsageValidation(workbook2)

    workbook2.save(generatedFile)        
    workbook2.close()


def SubUsageValidation(wb):    
    # wb = openpyxl.load_workbook(generatedFile )
    sheet1 = wb.get_sheet_by_name('Master Data')
    tab = Table(displayName="CommercialNonresidential", ref="S1:S26")
    sheet1.add_table(tab)
    tab = Table(displayName="IndustrialNonresidential", ref="T1:T5")
    sheet1.add_table(tab)
    tab = Table(displayName="InstitutionalNonresidential", ref="U1:U20")
    sheet1.add_table(tab)
    tab = Table(displayName="OthersNonresidential", ref="W1:W2")
    sheet1.add_table(tab)
    dest=wb.get_sheet_by_name('Property Assembly Detail')
    for index in range( 3,dest.max_row+1) :
        dv = DataValidation(type='list', formula1='=INDIRECT(SUBSTITUTE( SUBSTITUTE(SUBSTITUTE(H{0},"(",""),")",""), " ",""))'.format(index))
        dv.add('I{0}'.format(index))
        dest.add_data_validation(dv)
    # wb.save(generatedFile)
    # wb.close()

def addValidationToColumns (sheet, colName, formula,sheetName) :
    formula1 ="{0}!{1}".format(utils.quote_sheetname(sheetName),formula)
    dv = DataValidation(type="list", formula1=formula1 )
    # dv.error ='Your entry is not in the list'
    # dv.errorTitle =  'Invalid Entry'
    dv.add("{0}3:{1}{2}".format(colName,colName,sheet.max_row+1))
    sheet.add_data_validation(dv)
    return

def addValidationToColumns1 (sheet, colName, formula,sheetName) :
    formula1 ="{0}!{1}".format(utils.quote_sheetname(sheetName),formula)
    dv = DataValidation(type="list", formula1=formula1 )
    dv.add("{0}2:{1}{2}".format(colName,colName,1000))
    sheet.add_data_validation(dv)
    return


def getLocalityData(cityname):
    data = []
    boundary_path = os.path.join(config.MDMS_LOCATION ,  cityname , "egov-location")
    
    if os.path.isfile(os.path.join(boundary_path , "boundary-data.json")):
        with open(os.path.join(boundary_path , "boundary-data.json")) as f:
            existing_boundary_data = json.load(f)
    found = False
    for tenantboundary in existing_boundary_data["TenantBoundary"]:                
        for tenant_boundary in tenantboundary["boundary"]["children"]:
            for t1 in tenant_boundary["children"]:
                for t2 in t1["children"]:
                    data.append([t2["name"],t2["code"]])
                    #print(json.dumps(t1, indent = 2))
    dfLocality = pd.DataFrame(data,columns=['Locality Name','Code'])
    return dfLocality

def add_column(workbook,sheet_name, column):
    ws = workbook[sheet_name]
    new_column = ws.max_column + 1

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

    
if __name__ == "__main__":
    main()

    