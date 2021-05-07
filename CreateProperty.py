from common import *
from config import config, getValue, getMobileNumber, getTime,isna
import io
import os 
import sys
from common import superuser_login
from PropertyTax import *
import pandas as pd
import openpyxl
import collections
import re
from CreateWaterConnection import *
from CreateSewerageConnection import *
import traceback

now = datetime.now()
date_time = now.strftime("%d-%m-%Y") 

FOLDER_PATH  =r'D:\eGov\Data\WS\Azure Insertion'
# FOLDER_PATH  =r'C:\Users\Admin\Downloads\WaterSewerageTemplates'

def main() :
    print("Replace 109 of C:\ProgramData\Miniconda3\envs\py36\lib\site-packages\openpyxl\worksheet\merge.py with below one ") 
    print ("if side is None or  side.style is None:")
    root = FOLDER_PATH
    errorlogfile = open(os.path.join(root, "errorCBs.txt"), "w")  
    successlogfile = open(os.path.join(root, "CB With ProperData.txt"), "w")
    config.error_in_excel=[]
    config.DATA_ENTRY_ISSUES_FOLDER =os.path.join(root,date_time + '-Data_Entries_Issues')
    if not os.path.exists(config.DATA_ENTRY_ISSUES_FOLDER) :
        os.makedirs(config.DATA_ENTRY_ISSUES_FOLDER)
    if not os.path.exists(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR")) :
        os.makedirs(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR"))
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            cityname =module["code"].lower()[3:]
            config.errormsg=[]
            name = 'CB ' + cityname.lower()
            if  os.path.exists( os.path.join(root,name)):                
                try : 
                    if cityname =='nasirabad' : 
                        print("Processing for CB "+cityname.upper())
                        config.CITY_NAME = cityname
                        cbMain(cityname, successlogfile)
                except Exception as ex: 
                    print("Error in processing CB ",cityname , ex)
                    traceback.print_exc()
                    errorlogfile.write(cityname+"\n")
            if len(config.errormsg ) > 0 : 
                dateerror = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR",cityname+ "dateError.txt"), "w")  
                for element in config.errormsg:
                    dateerror.write(element + "\n") 
                dateerror.close()
    errorlogfile.close()
    successlogfile.close()
    if len(config.error_in_excel) > 0 : 
        cbHaveExcelIssue = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"_CB_HAVE_EXCEL_ISSUE.txt"), "w")  
        for element in config.error_in_excel:
                    cbHaveExcelIssue.write(element + "\n") 
        cbHaveExcelIssue.close()


def cbMain(cityname, successlogfile):
    Flag =False
    tenantMapping={}
    count = 0
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            tenantMapping[module["code"].lower()]=module["code"].lower()[3:]


    # Doing for one cb at a time
    # cityname = 'wellington'
    root = FOLDER_PATH
    name = 'CB ' + cityname.lower()
    propertyFile =os.path.join(root, name,'Template for Existing Property-Integrated with ABAS-' + cityname + '.xlsx')
    waterFile = os.path.join(root, name, "Template for Existing Water Connection Detail.xlsx")
    sewerageFile = os.path.join(root, name, "Template for Existing Sewerage Connection Detail.xlsx")
    logfile = open(os.path.join(root, name, "Logfile.json"), "w")   
    logfile.write("[ ")
    property_owner_obj = {}
    property_owner_obj = createOwnerObj(propertyFile)  
    if config.INSERT_DATA :
        validate = enterDefaultMobileNo(propertyFile, tenantMapping, cityname, waterFile, sewerageFile,logfile, property_owner_obj) 
        if(validate == False):                
            print('Data validation Failed for mobile entry, Please check the log file.') 
            return
    if os.path.exists(propertyFile) : 
        localityDict = getLocalityData(cityname) 
        validate =  validateDataForProperty(propertyFile, logfile,localityDict, cityname)
        if(validate == False):                
            print('Data validation for property Failed, Please check the log file.') 
            if config.INSERT_DATA: 
                return
        else:
            print('Data validation for property success.')                
        wb_property = openpyxl.load_workbook(propertyFile) 
        sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')         
        sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')
        
        df = pd.read_excel(propertyFile, sheet_name='Locality', usecols=['Locality Name', 'Code'])
        locality_data = {}
        for ind in df.index: 
            locality_data[df['Locality Name'][ind]] =  df['Code'][ind]   
        if config.INSERT_DATA and config.CREATE_PROPERTY: 
            createPropertyJson(sheet1, sheet2, locality_data,cityname, logfile, root, name)                
            wb_property.save(propertyFile)        
        wb_property.close()
    else:
        print("Property File doesnot exist for ", cityname) 
    
    if os.path.exists(waterFile) : 
        ProcessWaterConnection(propertyFile, waterFile, logfile, root, name,  cityname, property_owner_obj)          
    else:
        print("Water File doesnot exist for ", cityname) 

    if os.path.exists(sewerageFile) : 
        ProcessSewerageConnection(propertyFile, sewerageFile, logfile, root, name,  cityname, property_owner_obj)  
    else:
        print("Sewerage File doesnot exist for ", cityname) 

    logfile.seek(logfile.tell() - 1, os.SEEK_SET)
    logfile.write('')
    logfile.write("]")        

    size = os.path.getsize(os.path.join(root, name, "Logfile.json")) 
    logfile.close() 
    try :       
        
        if size > 2 : 
            df = pd.read_json (os.path.join(root, name, "Logfile.json"))
            
            df.to_excel(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER ,    name+ " Data Entries Issues.xlsx"), index = None)
            #df.to_excel(os.path.join(root, "WS_Data_Entry_Issues","CB "+ cityname+ " - Data Entries Issues.xlsx"), index = None)
            #df.to_csv (os.path.join(root, name, "DataValidation.csv"), index = None)
        else : 
            successlogfile.write("\n")
            successlogfile.write(cityname)
            successlogfile.write("\n")            
    except Exception as ex: 
        print("Error in parsing json file",ex)


def validateDataForProperty(propertyFile, logfile, localityDict, cityname):
    validated = True
    reason = ''

    try:
        wb_property = openpyxl.load_workbook(propertyFile) 
        sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')   
        sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')
        abas_ids = []        
        abas_ids_sheet2 = []        
        reason = 'Property file validation starts.\n'
        # validated = ValidateCols(logfile, propertyFile, sheet1, sheet2)
        # if not validated :
        #     print("Column Mismatch, sheets needs to be corrected")
        #     config["error_in_excel"].append(cityname +" have column issue in property sheet")

        # print('no. of rows in Property file sheet 1: ', sheet2.max_row ) 
        for index in range(2, sheet2.max_row +1): 
            if pd.isna(sheet2['A{0}'.format(index)].value):                    
                break
            propSheetABASId = getValue(sheet2['A{0}'.format(index)].value , str, '')           
            abas_ids_sheet2.append(propSheetABASId)

        emptyRows=0
        count =0 
        for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True): 
            try : 
                
                # if emptyRows > 10 :
                #     break
                if pd.isna(row[1]) and pd.isna(row[2]):
                    emptyRows =emptyRows +1
                    continue
                if pd.isna(row[0]):
                    validated = False
                    reason = 'Sl no. column is empty'
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),reason, getValue(row[1], str, ''))
                    #logfile.write(reason)
                if pd.isna(row[1]):
                    validated = False
                    #print("Property ID ",row[1])
                    reason = 'Property File sheet1 data validation failed for sl no. '+ getValue(row[0], str, '') + ', abas property id  is empty.\n'
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'abas property id  is empty',getValue(row[1], str, ''))
                    #logfile.write(reason)
                if pd.isna(row[27]):
                    validated = False
                    reason = 'Property File sheet1 data validation failed for sl no. '+ getValue(row[0], str, '') + ',  ownership type is empty.\n'
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'ownership type is empty',getValue(row[1], str, ''))
                    #logfile.write(reason)
                locality =getValue(row[13],str,"")
                if len ( locality) ==0:
                    validated = False
                    reason = 'Property File sheet1 data validation failed for sl no. '+ getValue(row[0], str, '') + ',  Locality/ Mohalla is empty.\n'
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'Locality/ Mohalla is empty',getValue(row[1], str, ''))
                    #logfile.write(reason)
                elif locality.lower() not in localityDict :
                    validated = False
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''), str(locality) +' Locality/ Mohalla does not exist in system ',getValue(row[1], str, ''))
                
                if(str(row[27]) != "Multiple Owners"):
                    if pd.isna(row[28]):
                        validated = False
                        reason = 'Property File sheet1 data validation failed for sl no. '+ getValue(row[0], str, '') + ', name is empty.\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'name is empty',getValue(row[1], str, ''))
                        #logfile.write(reason)
                    # elif not pd.isna(row[28]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[28]))):                    
                    #     validated = False
                    #     reason = 'Name has invalid characters for sl no. '+ getValue(row[0], int, '') +'\n'
                    #     write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'Name has invalid characters',getValue(row[], str, ''))
                    #     #logfile.write(reason)
                    if config.INSERT_DATA  and pd.isna(row[29]):
                        validated = False
                        reason = 'Property File data sheet1 validation failed for sl no. '+ str(getValue(row[0], str, '')) + ', mobile no. is empty.\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'mobile no. is empty',getValue(row[1], str, ''))
                        #logfile.write(reason)
                    elif not pd.isna(row[29])  and  not bool(re.match(config.MOBILE_PATTERN, getMobileNumber(row[29],str,""))) :
                        validated = False
                        count = count + 1
                        reason = 'Property File data sheet1 validation failed, Mobile number not correct for sl no. '+ getValue(row[0], str, '') +'\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'Mobile number not correct',getValue(row[1], str, ''))
                        #logfile.write(reason)
                    # if not pd.isna(row[41])  and  ( len(getMobileNumber(row[41],str,"")) != 11):
                    #     validated = False
                    #     reason = 'Property File data sheet1 validation failed, landline number not correct for sl no. '+ getValue(row[0], str, '') +'\n'
                    #     write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'landline number not correct',getValue(row[1], str, ''))
                    # if not pd.isna(row[33]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[33]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Guardian Name has invalid characters for abas id '+ getValue(row[0], int, '') +'\n'
                    #     write(logfile,propertyFile,sheet1.title,None,'Guardian Name has invalid characters',getValue(row[0], int, ''))
                    #     #logfile.write(reason)
                    if len(getValue(row[30], str, "")) > 0 and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[30]))) :                     
                        validated = False
                        reason = 'Property File data validation failed, Email id is not proper for abas id '+ getValue(row[1], str, '') +'\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'Email id is not proper',getValue(row[1], str, ''))
                        #logfile.write(reason)
                    if not pd.isna(row[32]) and getTime(row[32]) is None:
                        validated = False
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),str(row[32]) +' Invalid DOB format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))    
                
                elif(str(row[27]) == "Multiple Owners"):
                    propSheetABASId = getValue(row[1], str, "")
                    if propSheetABASId not in abas_ids_sheet2:
                        validated = False
                        reason = 'Property File data validation failed, abas id for multiple ownership is not available in Property Ownership Details sheet  '+ getValue(row[1], str, '') +'\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'abas id for multiple ownership is not available in Property Ownership Details sheet ',propSheetABASId)
                        #logfile.write(reason)
                propUsgType=getValue(row[7], str, "")
                if pd.isna(row[7]):
                    validated = False
                    reason = 'Property File data validation failed for sl no. '+ getValue(row[0], str, '') + ', usage type is empty.\n'
                    write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'usage type is empty',getValue(row[1], str, ''))
                    #logfile.write(reason)                    
                elif process_usage_type(propUsgType,True) is None:
                    validated = False
                    reason = 'Property File data validation failed for sl no. '+ getValue(row[0], str, '') + ', usage type is not correct.\n'
                    write(logfile,propertyFile,sheet1.title,getValue(row[7], str, ''),'usage type is not correct', getValue(row[1], str, ''))
                    #logfile.write(reason)    
                elif propUsgType.find("(") != -1 and propUsgType != 'Nonresidential ( Nonresidential )' :
                    subUsageValue =getValue(row[8], str, '')
                    if pd.isna(row[8]):
                        validated = False
                        reason = 'Property File data validation failed for sl no. '+ getValue(row[0], str, '') + ', sub usage type is empty.\n'
                        write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'sub usage type is empty',getValue(row[1], str, ''))
                        #logfile.write(reason)
                    elif subUsageValue.lower().strip() not in  USAGE_SUB_USAGE_MAP[process_usage_type(propUsgType,True)] :
                            validated = False            
                            # print("useage type ",subUsageValue.lower().strip())    
                            # print(USAGE_SUB_USAGE_MAP[process_usage_type(propUsgType,True)])        
                            write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),'sub usage "'+str(subUsageValue)+'" not correct as per usage type '+propUsgType,getValue(row[1], str, ''))
                            #raise Exception("dddd")

            except Exception as ex:
                print(config.CITY_NAME," validateDataForProperty Exception: ",getValue(row[0], int, ''), '  ',ex)
                # write(logfile,propertyFile,sheet1.title,getValue(row[0], int, ''),str(ex) ,getValue(row[1], str, ''))

        for index in range(3, sheet1.max_row +1):
            try: 
                if pd.isna(sheet1['B{0}'.format(index)].value):                    
                    break
                propSheetABASId = getValue(sheet1['B{0}'.format(index)].value, str, '')
                abas_ids.append(propSheetABASId)
            except Exception as ex:
                print( config.CITY_NAME,  " validateDataForProperty Exception: abas id is empty: ",getValue(row[0], int, ''), '  ',ex)
        duplicate_ids = [item for item, count in collections.Counter(abas_ids).items() if count > 1]

        if(len(duplicate_ids) >= 1):
            validated = False
            reason = 'Property File data validation failed. ' +'Duplicate abas property id for '+ str(duplicate_ids) +'\n'
            # print(reason)
            write(logfile,propertyFile,sheet1.title,None,'Duplicate ABAS property id for '+ str(duplicate_ids))
            #logfile.write(reason)      
        
        sheet2_index =1
        for row in sheet2.iter_rows(min_row=2, max_col=12, max_row=sheet2.max_row ,values_only=True):
            try :
                sheet2_index = sheet2_index +1
                abasId = getValue(row[0], str, '')
                ownerName = getValue(row[2], str, '')
                if  len(abasId)==0  and len(ownerName) > 0   : 
                     write(logfile,propertyFile,sheet2.title,sheet2_index,'ABAS ID is empty ',None)
                validateMobile =False      
                if not pd.isna(row[0]):
                    if validateMobile and pd.isna(row[3]):
                        validated = False
                        reason = 'Property File data validation failed for abas id  '+ getValue(row[0], str, '') + ', mobile no. is empty in multiple owner sheet.\n'
                        write(logfile,propertyFile,sheet2.title,None,'mobile no. is empty',getValue(row[0], str, ''))
                        #logfile.write(reason) 
                    elif not pd.isna(row[3]) and not bool(re.match(config.MOBILE_PATTERN, getMobileNumber(row[3],str,""))) :
                        validated = False
                        reason = 'Property File data validation failed, Mobile number not correct for abas id '+ getValue(row[0], str, '') +'\n'
                        write(logfile,propertyFile,sheet2.title,None,'Mobile number not correct',getValue(row[0], str, ''))
                        #logfile.write(reason)
                    if pd.isna(row[2]):
                        validated = False
                        reason = 'Property File data validation failed for abas id  '+ getValue(row[0], str, '') + ', name is empty.\n'
                        write(logfile,propertyFile,sheet2.title,None,'name is empty',getValue(row[0], str, ''))
                        #logfile.write(reason)
                    
                    # elif not pd.isna(row[2]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[2]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Name has invalid characters for abas id '+ getValue(row[0], str, '') +'\n'
                    #     write(logfile,propertyFile,sheet2.title,None,'Name has invalid characters',getValue(row[0], int, ''))
                    #     #logfile.write(reason)
                    # if not pd.isna(row[7]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[7]))):                        
                    #     validated = False
                    #     reason = 'Property File data validation failed, Guardian Name has invalid characters for abas id '+ getValue(row[0], str, '') +'\n'
                    #     write(logfile,propertyFile,sheet2.title,None,'Guardian Name has invalid characters',getValue(row[0], int, ''))
                    #     #logfile.write(reason)
                    if not pd.isna(row[4]) and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[4]))):                        
                        validated = False
                        reason = 'Property File data validation failed, Email id is not proper for abas id '+ str(getValue(row[0], str, '')) +'\n'
                        write(logfile,propertyFile,sheet2.title,None,'Email id is not proper',getValue(row[0], str, ''))
                        #logfile.write(reason)
                    if not isna(row[6]) and getTime(row[6]) is None : 
                        validated = False
                        write(logfile,propertyFile,sheet2.title,None, str(row[6]) +' Invalid DOB format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[0], str, ''))
  
            except Exception as ex:
                print(config.CITY_NAME," validateDataForProperty Exception: ",getValue(row[0], str, ''), '  ',ex)
                traceback.print_exc()
                # write(logfile,propertyFile,sheet2.title,None,str(ex) ,getValue(row[0], str, ''))
    except Exception as ex:
        print(config.CITY_NAME," validateDataForProperty Exception: ",ex)
        traceback.print_exc()
         
    reason = 'Property file validation ends.\n'
    #print(reason)
    # logfile.write(reason)
    return validated

def ValidateCols(logfile, propertyFile, sheet1, sheet2):
    proper_column_order1 = ['Sl No.', 'Existing Property ID* ( Unique Value on which property are getting searched in existing system ) ', 
    'Old Property code*', 'Property Type *', 'Total Land Area (in sq feet)\u2009 *', 'Total constructed Area (in sq feet)*', 
    'Existing Usage Type', 'Usage type *', 'Sub Usage type *', 'Occupancy Type*', 'Number of Floors*', 'Number of Flats ', 'CB Name *', 
    'Locality / Mohalla *\u2009\n', 'Ward No.', 'Block No.', 'Street Name', 'Building / Colony Name', 'House / Door No', 'Pin Code', 
    'Location of the Property *', 'ARV', 'RV', 'Financial Year', 'Is Property on Dispute(Yes/No) ', 'Is Property Authorized(Yes/No) ', 
    'Is Property  Encroached(Yes/No) ', 'OwnerShip Type*', 'Name/ Authorized Person*', 'Mobile No.*', 'Email Id', 'Gender*', 'DOB', 'Guardian Name*', 
    'Relationship*', 'Is correspondence address, same As Property Address*', 'Correspondence Address*', 'Special category*', 'Institution Name', 
    'Institution Type', 'Designation', 'Landline No']

    proper_column_order2 =  ['Existing Property ID*', 'OwnerShip Type*', 'Name*', 'Mobile No.*', 'Email Id', 'Gender*', 'DOB', 'Guardian Name*', 
    'Relationship*', 'Is correspondence address, same As Property Adderss*', 'Correspondence Address*', 'Special category*']
    validated = True
    column_list = [c.value for c in next(sheet1.iter_rows(min_row=2, max_row=2))]
    try:
        for i in range(0, 42):
            if(proper_column_order1[i].strip() != column_list[i].strip()) :
                print('Property file', column_list[i])
                validated = False
                write(logfile,propertyFile,sheet1.title,None,'Column order / name is not correct',column_list[i])
                # break

        column_list = [c.value for c in next(sheet2.iter_rows(min_row=1, max_row=1))]
        # print('\n')
        for i in range(0, 12):
            if(proper_column_order2[i].strip() != column_list[i].strip()) :
                print('Property file', column_list[i])
                validated = False
                write(logfile,propertyFile,sheet2.title,None,'Column order / name is not correct',column_list[i])
                # break
    except Exception as ex:
        validated = False
        print(config.CITY_NAME," validateCols Property Exception: ",ex)
        traceback.print_exc()
    return validated  

def createOwnerObj(propertyFile) :
    owner_obj = {}
    try:
        if os.path.exists(propertyFile) : 
            wb_property = openpyxl.load_workbook(propertyFile) 
            sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')        
            for i in range(3, sheet1.max_row +1):
                #print('B{0}'.format(i))
                if pd.isna(getValue(sheet1['B{0}'.format(i)].value, str, None)):                    
                    continue
                abas_id = getValue(sheet1['B{0}'.format(i)].value, str, '')
                for row in sheet1.iter_rows(min_row=i, max_col=42, max_row=i,values_only=True):                    
                    owner = {}               
                    owner['mobileNumber'] = getMobileNumber(row[29],str,"")
                    owner['ownerType'] =  getValue(row[27],str,"") 
                    if abas_id not in owner_obj:
                        owner_obj[abas_id] = []
                    owner_obj[abas_id].append(owner)       
            wb_property.close()
            return owner_obj
    except Exception as ex:
        print(config.CITY_NAME," createOwnerObj Exception: ",ex)
        traceback.print_exc()
    

def enterDefaultMobileNo(propertyFile, tenantMapping, cityname, waterFile, sewerageFile, logfile, owner_obj):
    validated = True
    search_key = 'pb.'+ cityname
    res = list(tenantMapping.keys()).index(search_key)
    res = res+1
    res = res* 100000
    mobileNumber = 3000000000 + res + 0
    # mobileNumber = 3002401400
    try:
        if os.path.exists(propertyFile) : 
            wb_property = openpyxl.load_workbook(propertyFile) 
            sheet1 = wb_property.get_sheet_by_name('Property Assembly Detail')   
            sheet2 = wb_property.get_sheet_by_name('Property Ownership Details')  
            index = 2
            for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True):
                index = index + 1
                if pd.isna(row[1]):
                    continue 
                if (str(row[27]).strip() != 'Multiple Owners'):
                    if pd.isna(row[29]):
                        mobileNumber = mobileNumber + 1                    
                        value = 'AD{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                        # logfile.write(value)
                        sheet1['AD{0}'.format(index)].value = mobileNumber
            index = 1
            for row in sheet2.iter_rows(min_row=2, max_col=5, max_row=sheet2.max_row ,values_only=True): 
                index = index + 1 
                if pd.isna(row[0]) :  
                    if (not pd.isna(row[1]) or not pd.isna(row[2])):
                        print("empty abas id in property file sheet2 while property validation")
                        write(logfile,propertyFile,sheet2.title,None,"empty abas id in property file sheet2 while property validation",None)
                        continue 
                    else : 
                        continue
                
                if pd.isna(row[3]):
                    mobileNumber = mobileNumber + 1
                    value = 'D{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                    # logfile.write(value)
                    # print(value)
                    sheet2['D{0}'.format(index)].value = mobileNumber
            wb_property.save(propertyFile)        
            wb_property.close()
        # else:
        #     print("Property File doesnot exist for ", cityname)  
        if os.path.exists(waterFile) :
            wb_water = openpyxl.load_workbook(waterFile) 
            water_sheet = wb_water.get_sheet_by_name('Water Connection Details') 
            index = 2
            for row in water_sheet.iter_rows(min_row=3, max_col=5, max_row=water_sheet.max_row ,values_only=True):
                index = index + 1
                if pd.isna(row[0]):
                    continue
                if(str(row[3]).strip().lower() == 'no'):
                    if pd.isna(row[4]):
                        mobileNumber = mobileNumber + 1
                        value = 'E{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                        logfile.write(value)
                        water_sheet['E{0}'.format(index)].value = mobileNumber
                
            wb_water.save(waterFile)        
            wb_water.close()
            # wb_water = openpyxl.load_workbook(waterFile) 
            # water_sheet = wb_water.get_sheet_by_name('Water Connection Details')
            # for row in water_sheet.iter_rows(min_row=3, max_col=5, max_row=water_sheet.max_row ,values_only=True):
            #     if pd.isna(getValue(row[1],str,None)):
            #         continue
            #     if(str(row[3]).strip().lower() == 'yes'):
            #         for obj in owner_obj[getValue(row[1],str,"")]:
            #             if(len(getMobileNumber(obj['mobileNumber'],str,"")) == 0):
            #                 validated = False
            #                 reason = 'Mobile number in property is not available as in water template same as owner detail for abas id. '+ str(row[1]).strip()
            #                 print(reason)
            #                 # logfile.write(reason)
        # else:
        #     print("Water File doesnot exist for ", cityname)  
        if os.path.exists(sewerageFile) :
            wb_sewerage = openpyxl.load_workbook(sewerageFile) 
            sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
            index = 2
            for row in sewerage_sheet.iter_rows(min_row=3, max_col=5, max_row=sewerage_sheet.max_row ,values_only=True):
                index = index + 1
                if pd.isna(row[0]):
                    continue
                if(str(row[3]).strip().lower() == 'no'):
                    if pd.isna(row[4]):
                        mobileNumber = mobileNumber + 1
                        value = 'E{0}'.format(index) + '    ' +str(mobileNumber) + '\n'
                        logfile.write(value)
                        sewerage_sheet['E{0}'.format(index)].value = mobileNumber
                
            wb_sewerage.save(sewerageFile)        
            wb_sewerage.close()
            # wb_sewerage = openpyxl.load_workbook(sewerageFile) 
            # sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
            # for row in sewerage_sheet.iter_rows(min_row=3, max_col=10, max_row=sewerage_sheet.max_row ,values_only=True):
            #     if pd.isna(getValue(row[1],str,None)):
            #         continue
            #     if(str(row[3]).strip().lower() == 'yes'):
            #         for obj in owner_obj[getValue(row[1],str,"")]:
            #             if(len(getMobileNumber(obj['mobileNumber'],str,"")) == 0):
            #                 validated = False
            #                 reason = 'Mobile number in property is not available as in sewerage template same as owner detail for abas id. '+ str(row[1]).strip()
            #                 # logfile.write(reason)

        # else:
        #     print("Sewerage File doesnot exist for ", cityname) 
        print("Last mobile number added: ", mobileNumber)
    except Exception as ex:
        print(config.CITY_NAME," DefaultMobileNo Exception: ",ex)
        traceback.print_exc()

    return validated

def createPropertyJson(sheet1, sheet2, locality_data,cityname, logfile,root, name):
    # abas_ids_multiple_owner = []
    # for index in range(2, sheet2.max_row +1):
    #     abas_ids_multiple_owner.append(sheet2['A{0}'.format(index)].value)
    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0
    auth_token = superuser_login()["access_token"]
    multiple_owner_obj = {}
    for i in range(2, sheet2.max_row +1):   
        try:
            if pd.isna(getValue(sheet2['A{0}'.format(i)].value, str, None)):                    
                continue     
            abas_id = getValue(sheet2['A{0}'.format(i)].value, str, None)
            for row in sheet2.iter_rows(min_row=i, max_col=12, max_row=i,values_only=True):                    
                owner = Owner()
                Owner.status = 'ACTIVE'
                owner.name = getValue( row[2] ,str,"NAME")
                owner.mobileNumber =  getMobileNumber( row[3] ,str,"3000000000")
                owner.emailId = getValue( row[4] ,str,"")
                owner.gender = process_gender(row[5])
                if not pd.isna(row[6]):
                    owner.dob = getTime(row[6])
                owner.fatherOrHusbandName = getValue( row[7] ,str,"Guardian")
                owner.relationship =  process_relationship(row[8])
                if getValue( row[9] ,str,"Yes").lower() == "yes":
                    owner.sameAsPeropertyAddress = True   
                else:
                    owner.sameAsPeropertyAddress = False 
                owner.ownerType =  process_special_category(str(row[11]).strip())
                if abas_id not in multiple_owner_obj:
                    multiple_owner_obj[abas_id] = []
                multiple_owner_obj[abas_id].append(owner)
        except Exception as ex:
            print(config.CITY_NAME," createPropertyJson Exception: ",ex)
            traceback.print_exc()
    index = 2
    for row in sheet1.iter_rows(min_row=3, max_col=42, max_row=sheet1.max_row ,values_only=True): 
                       
            index = index + 1  
            property = Property()  
            property.abasPropertyId =  getValue( row[1] ,str,None)
            if pd.isna(property.abasPropertyId):
                print("empty Abas id in property file for sl no. ", row[0])
                continue     
            
            locality = Locality()
            address  = Address()
            owner = Owner()
            unit = Unit()
            constructionDetail = ConstructionDetail()
            institution = Institution()
            additionalDetail = AdditionalDetails()
            tenantId = 'pb.'+ cityname
            property.tenantId = tenantId            
            
            status, res = property.search_abas_property(auth_token, tenantId, property.abasPropertyId)
            # with io.open(os.path.join(root, name,"property_search_res.json"), mode="w", encoding="utf-8") as f:
            #     json.dump(res, f, indent=2,  ensure_ascii=False)
            if(len(res['Properties']) == 0):  
                try:
                    print(cityname, "Property",property.abasPropertyId)
                    property.oldPropertyId =  getValue( row[2] ,str,None)
                    property.propertyType = process_property_type(str(row[3]).strip())            
                    property.landArea = getValue(row[4],float,1) 
                    property.superBuiltUpArea = getValue(row[5],float,1) 
                    if(int(property.landArea) == 0):
                        property.landArea = getValue(1,float,1)
                    if(int(property.superBuiltUpArea) == 0):
                        property.superBuiltUpArea = getValue(1,float,1)
                    property.usageCategory = process_usage_type(str(row[7]).strip())
                    if not (property.usageCategory == "RESIDENTIAL" or property.usageCategory == "MIXED" or property.usageCategory == "SLUM"
                            or property.usageCategory == "NONRESIDENTIAL.NONRESIDENTIAL") :
                        if(property.usageCategory == "NONRESIDENTIAL.OTHERS"):
                            property.subUsageCategory = "NONRESIDENTIAL.OTHERS.PUBLICFACILITY.CREMATIONBURIAL"
                        else:
                            property.subUsageCategory = process_sub_usage_type(str(row[8]).strip())     
                    else:
                        property.subUsageCategory = ''
                    if pd.isna(row[13]):
                        locality.code = "LOCAL_OTHERS"   
                    else:    
                        locality.code = locality_data[getValue(row[13] ,str,"Others")]
                    address.city = cityname
                    address.locality = locality
                    address.location = process_location(str(row[20]).strip())
                    address.street = getValue(row[16] ,str,None)
                    address.buildingName = getValue(row[17] ,str,"")
                    address.doorNo = getValue(row[18],str,"")
                    address.pincode = getValue(row[19],str,None)
                    correspondence_address = get_propertyaddress(address.doorNo,address.buildingName,getValue( row[13] ,str,"Others"),cityname)
                    unit.occupancyType = process_occupancy_type(str(row[9]).strip())
                    unit.arv = getValue(row[21],int,0) 
                    unit.floorNo = 0
                    if(len(property.subUsageCategory) == 0):
                        unit.usageCategory = property.usageCategory
                    else:                
                        unit.usageCategory = property.subUsageCategory
                    constructionDetail.builtUpArea = property.superBuiltUpArea 
                    unit.construction_detail = constructionDetail        
                    property.address = address
                    property.units = []
                    property.units.append(unit)
                    property.owners = []
                    # converter = lambda  x,y  : x  if x is not pd.isna else y
                    property.noOfFloors = getValue(row[10],int,1) 
                    if(int(property.noOfFloors) == 0):
                        property.noOfFloors = getValue(1,int,1)
                    property.noOfFlats = getValue(row[11],int,0) 
                    financial_year = getValue(row[23],str,"2020-2021").replace("-20", "-")
                    property.financialYear = financial_year
                    property.ownershipCategory = process_ownership_type(str(row[27]).strip())     
                    if(property.ownershipCategory == 'INDIVIDUAL.SINGLEOWNER'):
                        owner.status = 'ACTIVE'            
                        owner.name = getValue(row[28] ,str,"NAME")
                        owner.mobileNumber = getValue(row[29],str,"3000000000")
                        owner.emailId = getValue(row[30] ,str,"")
                        owner.gender = process_gender(row[31] )
                        if not pd.isna(row[32]):
                            owner.dob = getTime(row[32])
                        owner.fatherOrHusbandName = getValue(row[33],str,"Guardian")
                        owner.relationship =  process_relationship(row[34])
                        if getValue( row[35] ,str,"Yes").lower() == "yes":
                            owner.sameAsPeropertyAddress = True   
                        else:
                            owner.sameAsPeropertyAddress = False 
                        if(owner.sameAsPeropertyAddress ==  True):
                            owner.correspondenceAddress = correspondence_address
                        else: 
                            owner.correspondenceAddress = getValue(row[36],str,correspondence_address)
                        owner.ownerType =  process_special_category(str(row[37]).strip())
                        
                        property.owners.append(owner)
                    elif(property.ownershipCategory == 'INSTITUTIONALPRIVATE'):
                        owner.status = 'ACTIVE'
                        owner.name = getValue(row[28],str,"NAME")
                        owner.mobileNumber = getValue(row[29],str,"3000000000")
                        owner.emailId = getValue(row[30],str,"")
                        if getValue( row[35] ,str,"Yes").lower() == "yes":
                            owner.sameAsPeropertyAddress = True   
                        else:
                            owner.sameAsPeropertyAddress = False 
                        institution.name = getValue(row[38],str,"Institution")
                        institution.type = process_private_institution_type(str(row[39]).strip())
                        institution.designation = getValue(row[40],str,"Designation")
                        owner.altContactNumber = getValue(row[41],str,"10000000000")
                        if(owner.sameAsPeropertyAddress ==  True):
                            owner.correspondenceAddress = correspondence_address
                        else: 
                            owner.correspondenceAddress = getValue(row[36],str,"Correspondence")
                        owner.ownerType =  process_special_category(str(row[37]).strip())    
                        property.institution = institution
                        property.owners.append(owner)
                    elif(property.ownershipCategory == 'INSTITUTIONALGOVERNMENT'):
                        owner.status = 'ACTIVE'
                        owner.name = getValue(row[28],str,"NAME")
                        owner.mobileNumber = getValue(row[29],str,"3000000000")
                        owner.emailId = getValue(row[30],str,"")
                        if getValue( row[35] ,str,"Yes").lower() == "yes":
                            owner.sameAsPeropertyAddress = True   
                        else:
                            owner.sameAsPeropertyAddress = False 
                        institution.name = getValue(row[38],str,"Institution")
                        institution.type = process_govt_institution_type(str(row[39]).strip())
                        institution.designation = getValue(row[40],str,"Designation")
                        owner.altContactNumber = getValue(row[41],str,"10000000000")
                        if(owner.sameAsPeropertyAddress ==  True):
                            owner.correspondenceAddress = correspondence_address
                        else: 
                            owner.correspondenceAddress = getValue(row[36],str,"Correspondence")
                        owner.ownerType =  process_special_category(str(row[37]).strip())    
                        property.institution = institution
                        property.owners.append(owner)
                    elif(property.ownershipCategory == 'INDIVIDUAL.MULTIPLEOWNERS'):
                        for owner_obj in multiple_owner_obj[property.abasPropertyId]:
                            owner = owner_obj
                            owner.status = 'ACTIVE'  
                            if(owner.sameAsPeropertyAddress ==  True):
                                owner.correspondenceAddress = correspondence_address
                            else: 
                                owner.correspondenceAddress = getValue(row[10],str,"Correspondence")
                            property.owners.append(owner)

                        # occurances = [i for i,x in enumerate(abas_ids_multiple_owner) if x == property.abas_property_id]
                        # for i in occurances:                
                        #     for row in sheet2.iter_rows(min_row=i+2, max_col=12, max_row=i+2,values_only=True):                    
                        #         owner = Owner()
                        #         owner.status = 'ACTIVE'
                        #         owner.name = getValue(str(row[2]).strip(),str,"NAME")
                        #         owner.mobileNumber = getValue(str(row[3]).strip(),str,"3000000001")
                        #         owner.emailId = getValue(str(row[4]).strip(),str,"")
                        #         owner.gender = process_gender(row[5])
                        #         owner.fatherOrHusbandName = getValue(str(row[7]).strip(),str,"Guardian")
                        #         owner.relationship =  process_relationship(row[8])
                        #         owner.sameAsPeropertyAddress = getValue(str(row[9]).strip(),bool,True)
                        #         if(owner.sameAsPeropertyAddress ==  True):
                        #             owner.correspondenceAddress = correspondence_address
                        #         else: 
                        #             owner.correspondenceAddress = getValue(str(row[10]).strip(),str,"Correspondence")
                        #         owner.ownerType =  process_special_category(str(row[11]).strip())
                        #         property.owners.append(owner)

                    additionalDetail.isRainwaterHarvesting = False
                    additionalDetail.isPropertyDisputed = process_YesNo(str(row[24]).strip())
                    additionalDetail.isPropertyAuthorized = process_YesNo(str(row[25]).strip())
                    property.additional_details= additionalDetail
                    property.source = 'LEGACY_RECORD'
                    property.channel = 'MIGRATION'
                    property.creationReason = 'DATA_UPLOAD'
                except Exception as ex:
                    print(config.CITY_NAME," createPropertyJson Exception: ",ex)
                    traceback.print_exc()
                # print('property ', property.get_property_json())            
                
                req_data,statusCode, res = property.upload_property(auth_token, tenantId, property.abasPropertyId,root, name)
                # with io.open(os.path.join(root, name,"property_create_res.json"), mode="w", encoding="utf-8") as f:
                #     json.dump(res, f, indent=2,  ensure_ascii=False)
                propertyId = ''
                if(statusCode == 200 or statusCode == 201):
                    for found_index, resProperty in enumerate(res["Properties"]):
                        propertyId = resProperty["propertyId"]
                        # value = 'B{0}'.format(index) + '    ' + str(propertyId) + '\n'
                        # logfile.write(value)
                        sheet1['AQ{0}'.format(index)].value = propertyId
                        reason = 'property created for abas id ' + str(property.abasPropertyId)
                        # logfile.write(reason)
                        # print(reason)
                        createdCount = createdCount + 1
                        break
                else:
                    with io.open(os.path.join(root, name, str(property.abasPropertyId) + "_property_create_req.json"), mode="w", encoding="utf-8") as f:
                        json.dump(req_data, f, indent=2,  ensure_ascii=False)
                    with io.open(os.path.join(root, name, str(property.abasPropertyId) + "_property_create_res.json"), mode="w", encoding="utf-8") as f:
                        json.dump(res, f, indent=2,  ensure_ascii=False)
                    reason = 'property not created status code '+ str(statusCode) + ' for abas id ' + str(property.abasPropertyId) + ' response: ', str(res)  + '\n'
                    # logfile.write(reason)
                    print(reason)
                    notCreatedCount = notCreatedCount + 1
            else:
                for found_index, resProperty in enumerate(res["Properties"]):
                    propertyId = resProperty["propertyId"]
                    break
                sheet1['AQ{0}'.format(index)].value = propertyId
                reason = 'property already exist for abas id '+ str(property.abasPropertyId) + '\n'
                # logfile.write(reason)
                # print(reason)
                searchedCount = searchedCount + 1

    reason = 'property created count: '+ str(createdCount)
    print(reason)
    reason = 'property not created count: '+ str(notCreatedCount)
    print(reason)
    reason = 'property searched count: '+ str(searchedCount)
    print(reason)


def get_propertyaddress(doorNo, buildingName,locality,cityname):
    return doorNo + ' ' + buildingName + ' ' +locality + ' ' + cityname

def process_YesNo(value):
    YesNo_MAP = {
        "Yes": True,
        "No": False,
        "None": None
    }
    return YesNo_MAP[value]

def process_location(value):
    location_MAP = {
        "Inside Civil Area (Bazar Area)": "CIVIL",
        "Outside Civil Area (Bungalow Area)": "BUNGLOW",
        "Outside Cantonment Area": "OUTSIDE",
        "None": "CIVIL"
    }
    return location_MAP[value]

def process_relationship(value):
    if value is None : 
        value ="parent"
    value = value.strip().lower()
    relationship_MAP = {
        "parent": "PARENT",
        "spouse": "SPOUSE",
        "gurdian": "GUARDIAN",
        "guardian": "GUARDIAN",
        "none": "PARENT",
        "na": "PARENT"
    }
    return relationship_MAP[value]

def process_gender(value):
    if value is None : 
        value ="Male"
    value =str(value).strip().lower()
    gender_MAP = {
        "male": "MALE",
        "female": "FEMALE",
        "transgender": "TRANSGENDER",
        "none": "MALE"       
    }
    return gender_MAP[value]

def process_private_institution_type(value):
    private_institution_MAP = {
        "NA": "OTHERSPRIVATEINSTITUITION",
        "None": "OTHERSPRIVATEINSTITUITION",
        "Private company": "PRIVATECOMPANY",
        "NGO": "NGO",
        "Private Trust": "PRIVATETRUST",
        "Private Board": "PRIVATEBOARD",
        "Other Private Institution": "OTHERSPRIVATEINSTITUITION"        
    }
    return private_institution_MAP[value]

def process_govt_institution_type(value):
    govt_institution_MAP = {
        "None": "OTHERGOVERNMENTINSTITUITION",
        "CB Government": "ULBGOVERNMENT",
        "state Government": "INSTITUTIONALGOVERNMENT",
        "Central Government": "CENTRALGOVERNMENT",
        "Other Government Institution": "OTHERGOVERNMENTINSTITUITION"
    }
    return govt_institution_MAP[value]

def process_property_type(value):
    PT_MAP = {
        "None": "BUILTUP.INDEPENDENTPROPERTY",
        "Vacant Land": "VACANT",
        "Flat/Part of the building": "BUILTUP.SHAREDPROPERTY",
        "Independent Building": "BUILTUP.INDEPENDENTPROPERTY"
    }
    return PT_MAP[value]

def process_occupancy_type(value):
    if value is None : 
        value ="Self-Occupied"
    value =str(value).strip().lower()
    OC_MAP = {
        "na": "SELFOCCUPIED",
        "residential": "SELFOCCUPIED",
        "none": "SELFOCCUPIED",
        "self-occupied": "SELFOCCUPIED",
        "selfoccupied": "SELFOCCUPIED",
        "rented": "RENTED",
        "unoccupied": "UNOCCUPIED"
    }
    return OC_MAP[value]


 
        

## As its static data so need to load it again
USAGE_MAP = {
        "Residential": "RESIDENTIAL",
        "Nonresidential ( Nonresidential )": "NONRESIDENTIAL.NONRESIDENTIAL",
        "Commercial ( Nonresidential )": "NONRESIDENTIAL.COMMERCIAL",
        "Industrial ( Nonresidential )": "NONRESIDENTIAL.INDUSTRIAL",
        "Institutional ( Nonresidential )": "NONRESIDENTIAL.INSTITUTIONAL",
        "Others ( Nonresidential )": "NONRESIDENTIAL.OTHERS",
        "Mixed": "MIXED",
        "Slum": "SLUM",
        "None": "RESIDENTIAL"
    }
USAGE_MAP = { k.strip().lower():USAGE_MAP[k] for k in USAGE_MAP}

def process_usage_type(value, isValidation =False):
    value =value.strip().lower()
    if isValidation : 
        if value in USAGE_MAP : 
            return USAGE_MAP[value.strip().lower()] 
        return None
    return USAGE_MAP[value]
SUB_USAGE_MAP = {
        'None':'' ,'Animal Dairy(Below 10 Cattle)':'NONRESIDENTIAL.COMMERCIAL.ANIMALDAIRYLESS' , 'Animal Dairy(Above 10 Cattle)':'NONRESIDENTIAL.COMMERCIAL.ANIMALDAIRYMORE' , 'Bank':'NONRESIDENTIAL.COMMERCIAL.BANK' , 'Dhobi':'NONRESIDENTIAL.COMMERCIAL.DHOBI' , 'Dyers':'NONRESIDENTIAL.COMMERCIAL.DYERS' , 'Movie Theatre':'NONRESIDENTIAL.COMMERCIAL.ENTERTAINMENT.MOVIETHEATRE' , 'Multiplex':'NONRESIDENTIAL.COMMERCIAL.ENTERTAINMENT.MULTIPLEX' , 'Marriage Palace':'NONRESIDENTIAL.COMMERCIAL.EVENTSPACE.MARRIAGEPALACE' , 'Ac Restaurant':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.ACRESTAURANT' , 'Non Ac Restaurant':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.NONACRESTAURANT' , 'Bhojanalaya/Tea Shop/Halwai Shop':'NONRESIDENTIAL.COMMERCIAL.FOODJOINTS.TEA' , 'Hotels':'NONRESIDENTIAL.COMMERCIAL.HOTELS' , 'Pathlab':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PATHLAB' , 'Private Dispensary':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PVTDISPENSARY' , 'Private Hospital':'NONRESIDENTIAL.COMMERCIAL.MEDICALFACILITY.PVTHOSPITAL' , 'Office Space(Less Than 10 Persons)':'NONRESIDENTIAL.COMMERCIAL.OFFICESPACELESS' , 'Office Space(More Than 10 Persons)':'NONRESIDENTIAL.COMMERCIAL.OFFICESPACEMORE' , 'Other Commercial Usage':'NONRESIDENTIAL.COMMERCIAL.OTHERCOMMERCIALSUBMINOR.OTHERCOMMERCIAL' , 'Petrol Pump':'NONRESIDENTIAL.COMMERCIAL.PETROLPUMP' , 'Grocery Store':'NONRESIDENTIAL.COMMERCIAL.RETAIL.GROCERY' , 'Malls':'NONRESIDENTIAL.COMMERCIAL.RETAIL.MALLS' , 'Pharmacy':'NONRESIDENTIAL.COMMERCIAL.RETAIL.PHARMACY' , 'Showroom':'NONRESIDENTIAL.COMMERCIAL.RETAIL.SHOWROOM' , 'Service Centre':'NONRESIDENTIAL.COMMERCIAL.SERVICECENTER' , 'Statutory Organisation':'NONRESIDENTIAL.COMMERCIAL.STATUTORY.STATUTORYORGANISATION' , 'Manufacturing Facility(Less Than 10 Persons)':'NONRESIDENTIAL.INDUSTRIAL.MANUFACTURINGFACILITY.MANUFACTURINGFACILITYLESS' , 'Manufacturing Facility(More Than 10 Persons)':'NONRESIDENTIAL.INDUSTRIAL.MANUFACTURINGFACILITY.MANUFACTURINGFACILITYMORE' , 'Other Industrial Usage':'NONRESIDENTIAL.INDUSTRIAL.OTHERINDUSTRIALSUBMINOR.OTHERINDUSTRIAL' , 'Godown/Warehouse':'NONRESIDENTIAL.INDUSTRIAL.WAREHOUSE.WAREHOUSE' , 'College':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.COLLEGES' , 'Other Private Educational Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.OTHEREDUCATIONAL' , 'Polytechnic':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.POLYTECHNICS' , 'School':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.SCHOOL' , 'Training Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONAL.TRAININGINSTITUTES' , 'Govt. Aided Educational Institute':'NONRESIDENTIAL.INSTITUTIONAL.EDUCATIONALGOVAIDED.GOVAIDEDEDUCATIONAL' , 'Historical Building':'NONRESIDENTIAL.INSTITUTIONAL.HISTORICAL.HISTORICAL' , 'Stray Animal Care Center':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.ANIMALCARE' , 'Home For The Disabled / Destitute':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.DISABLEDHOME' , 'Old Age Homes':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.OLDAGEHOMES' , 'Orphanage':'NONRESIDENTIAL.INSTITUTIONAL.HOMESFORSPECIALCARE.ORPHANAGE' , 'Others':'NONRESIDENTIAL.INSTITUTIONAL.OTHERINSTITUTIONALSUBMINOR.OTHERINSTITUTIONAL' , 'Community Hall':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.COMMUNITYHALL' , 'Govt. Hospital & Dispensary':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.GOVTHOSPITAL' , 'Public Libraries':'NONRESIDENTIAL.INSTITUTIONAL.PUBLICFACILITY.LIBRARIES' , 'Golf Club':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.GOLFCLUB' , 'Social Club':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.SOCIALCLUB' , 'Sports Stadium':'NONRESIDENTIAL.INSTITUTIONAL.RECREATIONAL.SPORTSSTADIUM' , 'Religious':'NONRESIDENTIAL.INSTITUTIONAL.RELIGIOUSINSTITUITION.RELIGIOUS' , 'Cremation/ Burial Ground':'NONRESIDENTIAL.OTHERS.PUBLICFACILITY.CREMATIONBURIAL'
}
SUB_USAGE_MAP = { k.strip().lower():SUB_USAGE_MAP[k] for k in SUB_USAGE_MAP}

USAGE_SUB_USAGE_MAP =dict()
for ele in USAGE_MAP :
    if USAGE_MAP[ele] not in USAGE_SUB_USAGE_MAP  : 
        USAGE_SUB_USAGE_MAP[USAGE_MAP[ele]]=list()
for sub_usg_key in SUB_USAGE_MAP : 
    sub_usg_val =SUB_USAGE_MAP[sub_usg_key]
    for ele in USAGE_SUB_USAGE_MAP :
        if sub_usg_val.startswith(ele) :
            USAGE_SUB_USAGE_MAP[ele].append(sub_usg_key.strip().lower())

def process_sub_usage_type(value,  isValidation =False ):  
    value =value.strip().lower()
    return SUB_USAGE_MAP[value]

def process_ownership_type(value):
    if value is None : 
        value ="None"
    value =value.strip().lower()
    Ownsership_MAP = {
        "none": "INDIVIDUAL.SINGLEOWNER",
        "single owner": "INDIVIDUAL.SINGLEOWNER",
        "multiple owners": "INDIVIDUAL.MULTIPLEOWNERS",
        "institutional- private": "INSTITUTIONALPRIVATE",
        "institutional- government": "INSTITUTIONALGOVERNMENT"
    }
    return Ownsership_MAP[value]

def process_special_category(value):
    if value is None: 
        value ="None"
    value =value.strip().lower()
    special_category_MAP = {
        "freedom fighter": "FREEDOMFIGHTER",
        "widow": "WIDOW",
        "handicapped": "HANDICAPPED",
        "below poverty line": "BPL",
        "defense personnel": "DEFENSE",
        "employee/staff of cb": "STAFF",
        "none of the above": "NONE",
        "none":"NONE",
        "na":"NONE"
    }
    return special_category_MAP[value]

def getLocalityData(cityname):
    data = []
    boundary_path = os.path.join(config.MDMS_LOCATION ,  cityname , "egov-location")
    localityDict =dict()
    if os.path.isfile(os.path.join(boundary_path , "boundary-data.json")):
        with open(os.path.join(boundary_path , "boundary-data.json")) as f:
            existing_boundary_data = json.load(f)
    found = False
    for tenantboundary in existing_boundary_data["TenantBoundary"]:                
        for tenant_boundary in tenantboundary["boundary"]["children"]:
            for t1 in tenant_boundary["children"]:
                for t2 in t1["children"]:
                    localityDict[t2["name"].lower()] =t2["code"]
    return localityDict


        

if __name__ == "__main__":
    # print(process_usage_type("Slum",True))
    main()
    
    
              


