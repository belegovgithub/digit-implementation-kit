from common import *
from config import config
import io
import os
import numpy
import pandas as pd
import openpyxl
import traceback
from config import getValue, isna


def main():
    filePath=r"D:\localization\ForUpdate_14May2021"
    fileName="tamil"
    # sheetName = "ws"
    # dfs = open_excel_file(os.path.join(filePath,fileName+".xlsx"))
    # df = get_sheet(dfs, sheetName)
    workBook = openpyxl.load_workbook(os.path.join(filePath,fileName+".xlsx"))
    sheets = workBook.sheetnames
    #sheets = ['tl','uc','pgr','ws','pt']
    count = 0
    ######For a particular sheet #################
    # try:
    #     df = pd.read_excel(os.path.join(filePath,fileName+".xlsx"), sheet_name=sheetName, usecols=['code', 'message'])
    #     print(sheetName)
    #     localization_data=[]
    #     for ind in df.index: 
    #         if not isna(getValue(df['code'][ind],str,'')) :
    #             localization_data.append({
    #                             "code":df['code'][ind]  ,
    #                             "message": df['message'][ind],
    #                             "module": "rainmaker-" + sheetName,
    #                             "locale": "ml_IN" 
    #                         })

    #     data = {
    #         "RequestInfo": {
    #             "authToken": "{{access_token}}"
    #         },
    #         "tenantId": config.TENANT,
    #         "messages": localization_data
    #     }

    #     auth_token = superuser_login()["access_token"]        
    #     localize_response = upsert_localization(auth_token, data)
    #     if(not (localize_response.status_code == 200 or localize_response.status_code == 201)):
    #         with io.open(os.path.join(filePath, fileName+ "_" +sheetName+ ".json"), mode="w", encoding="utf-8") as f:
    #             json.dump(data, f, indent=2,  ensure_ascii=False)
    #         print(localize_response.json())
    #         print("Upsert Failed")  
    #     else:
    #         count = count + 1
    #     print(count)  
    # except:
    #     traceback.print_exc() 
    # return
    ##########End of Paricular sheet localization#############
    ######For full excel file #################
    for sheetName in sheets:
        try:
            print(sheetName)
            df = pd.read_excel(os.path.join(filePath,fileName+".xlsx"), sheet_name=sheetName, usecols=['code', 'message'])
            
            localization_data=[]
            TENANT = "pb"
            if(sheetName.find(".") >0):
                TENANT = sheetName

            for ind in df.index: 
                if not isna(getValue(df['code'][ind],str,'')) :
                    localization_data.append({
                                    "code":df['code'][ind]  ,
                                    "message": df['message'][ind],
                                    "module": "rainmaker-" + sheetName.lower(),
                                    # "module": sheetName.lower(),
                                    "locale": "ta_IN" 
                                })
            
            data = {
                "RequestInfo": {
                    "authToken": "{{access_token}}"
                },
                "tenantId": TENANT,
                "messages": localization_data
            }            
        
            auth_token = superuser_login()["access_token"]        
            localize_response = upsert_localization(auth_token, data)
            if(not (localize_response.status_code == 200 or localize_response.status_code == 201)):
                with io.open(os.path.join(filePath, fileName+ "_" +sheetName+ ".json"), mode="w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2,  ensure_ascii=False)
                print(localize_response.json())
                print("Upsert Failed")  
            else:
                count = count + 1
        except:
            traceback.print_exc() 
    print(count) 

    ###############End of full excel file Localization#################################

def updateLocalization():
    filePath=r"C:\Users\Administrator\Downloads"
    fileName="TeleguUpdate_2203"
    # sheetName = "ws"
    # dfs = open_excel_file(os.path.join(filePath,fileName+".xlsx"))
    # df = get_sheet(dfs, sheetName)
    workBook = openpyxl.load_workbook(os.path.join(filePath,fileName+".xlsx"))
    sheets = workBook.sheetnames
    sheets = ['Sheet1']
    count = 0
    for sheetName in sheets:
        df = pd.read_excel(os.path.join(filePath,fileName+".xlsx"), sheet_name=sheetName, usecols=['code', 'message','module'])
        print(sheetName)
        localization_data=[]
        TENANT = "pb"
        if(sheetName.find(".") >0):
            TENANT = sheetName

        for ind in df.index: 
            localization_data.append({
                            "code":df['code'][ind]  ,
                            "message": df['message'][ind],
                            "module": df['module'][ind],
                            "locale": "te_IN" 
                        })
        
        data = {
            "RequestInfo": {
                "authToken": "{{access_token}}"
            },
            "tenantId": TENANT,
            "messages": localization_data
        }
        # with io.open(os.path.join(filePath, fileName+ ".json"), mode="w", encoding="utf-8") as f:
        #     json.dump(data, f, indent=2,  ensure_ascii=False)
    
        auth_token = superuser_login()["access_token"]        
        localize_response = upsert_localization(auth_token, data)
        if(not (localize_response.status_code == 200 or localize_response.status_code == 201)):
            print(localize_response.json())
            print("Upsert Failed")  
        else:
            count = count + 1
    print(count) 
    ###############End of full excel file Localization#################################

def process_CB_localization(CBNAME, district, district_code, state):
    locale_data = []
    locale_module = "rainmaker-common"
    locale_data.append({
                        "code": "TENANT_TENANTS_PB_"+ CBNAME,
                        "message": config.CITY_NAME,
                        "module": locale_module,
                        "locale": "en_IN"
                    })
    locale_data.append({
                        "code": "PB_"+ CBNAME + "_" + district_code + "_LABEL",
                        "message": district,
                        "module": locale_module,
                        "locale": "en_IN"
                    })
    locale_data.append({
                        "code": "MYCITY_"+ CBNAME + "_" + "STATE_LABEL",
                        "message": state,
                        "module": locale_module,
                        "locale": "en_IN"
                    })
    data = {
        "RequestInfo": {
            "authToken": "{{access_token}}"
        },
        "tenantId": config.TENANT,
        "messages": locale_data
    }
    auth_token = superuser_login()["access_token"]
    localize_response = upsert_localization(auth_token, data)
    print(localize_response)
    print("Tenant localization for english is pushed.")

if __name__ == "__main__":
    main()
    #updateLocalization()
    
