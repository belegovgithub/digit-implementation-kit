from common import *
from config import config, getValue, getMobileNumber, getTime, isna
import io
import os 
import sys
from WaterConnection import *
from PropertyTax import *
import pandas as pd
import openpyxl
import collections
import traceback
    
def ProcessWaterConnection(propertyFile, waterFile, logfile, root, name,  cityname, property_owner_obj = {}) :
    wb_property = openpyxl.load_workbook(propertyFile) 
    propertySheet = wb_property.get_sheet_by_name('Property Assembly Detail') 
    wb_water = openpyxl.load_workbook(waterFile) 
    waterSheet = wb_water.get_sheet_by_name('Water Connection Details')  
    #print('no. of rows in water file: ', waterSheet.max_row ) 
    validate = validateWaterData(propertySheet, waterFile, logfile, cityname, property_owner_obj)  
    if(validate == False):                
        print('Data validation for water Failed, Please check the log file.') 
        if config.INSERT_DATA: 
            return
    else:
        print('Data validation for water success.')
    if config.INSERT_DATA and config.CREATE_WATER: 
        createWaterJson(propertySheet, waterSheet, cityname, logfile, root, name)   
        wb_water.save(waterFile)        
    wb_water.close()

def validateWaterData(propertySheet, waterFile, logfile, cityname, property_owner_obj):
    validated = True
    wb_water = openpyxl.load_workbook(waterFile) 
    water_sheet = wb_water.get_sheet_by_name('Water Connection Details')     
    index = 2
    reason = 'Water file validation starts.\n'
    # validated = ValidateCols(waterFile, water_sheet, logfile)
    # if not validated :
    #     print("Column Mismatch, sheets needs to be corrected")
    #     config["error_in_excel"].append(cityname +" have column issue in water sheet")

    abas_ids = [] 
    old_connections = []
    try:
        for index in range(3, propertySheet.max_row +1 ):
            if pd.isna(propertySheet['B{0}'.format(index)].value):
                # validated = False
                # reason = 'Water File data validation failed, Sl no. column is empty\n'
                # write(logfile,"property excel",propertySheet.title,index,'Sl no. column is empty')
                #logfile.write(reason)
                continue
            propSheetABASId = getValue(propertySheet['B{0}'.format(index)].value, str,'')
            abas_ids.append(str(propSheetABASId).strip())     
    except Exception as ex:
        print(config.CITY_NAME," validateWaterData Exception: ", ex)  
 
    emptyRows =0 
    for row in water_sheet.iter_rows(min_row=3, max_col=26, max_row=water_sheet.max_row ,values_only=True):
        index = index + 1
        try:        
            # if emptyRows > 10 :
            #     break
            if pd.isna(row[1]) and pd.isna(row[2]):
                emptyRows = emptyRows +1
                continue
            if pd.isna(row[0]):
                validated = False
                reason = 'Sl no. column is empty\n'
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Sl no. column is empty',getValue(row[1], str, ''))
                #logfile.write(reason)
            if pd.isna(row[1]): #len(getValue(row[1], str, '')) == 0:
                validated = False
                reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', abas id is empty.\n'
                #logfile.write(reason) 
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'abas id is empty',getValue(row[1], str, ''))
            if pd.isna(row[2]):
                validated = False
                reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', existing water connection number is empty.\n'
                #logfile.write(reason)
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'existing water connection number is empty',getValue(row[1], str, ''))
            if pd.isna(row[3]):
                validated = False
                reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', same as property address cell is empty.\n'
                #logfile.write(reason)
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'same as property address cell is empty',getValue(row[1], str, ''))
            if getValue(row[16], str, '') == "Metered":
                if pd.isna(row[20]):
                    validated = False
                    reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', Meter Id is empty.\n'
                    #logfile.write(reason)
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Meter Id is empty',getValue(row[1], str, ''))
                if pd.isna(row[21]):
                    validated = False
                    reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', last meter reading is empty.\n'
                    #logfile.write(reason)
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'last meter reading is empty',getValue(row[1], str, ''))
                elif not bool(re.match("^[0-9]+$",getValue(row[21], str, ''))) :                      
                    validated = False
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Last meter reading must be a numeric value.',getValue(row[1], str, ''))
                    #logfile.write(reason)
                if pd.isna(row[22]):
                    validated = False
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'last billed date is empty',getValue(row[1], str, ''))
                elif pd.isna(getTime(row[22])):
                    validated = False
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),str(row[22]) +' Invalid last billed date format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))
            if isna(row[18]):
                    validated = False
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Activation date is empty',getValue(row[1], str, ''))
            elif getTime(row[18]) is None:
                validated = False
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),str(row[18]) +' Invalid Activation date format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))
            if(str(row[3]).strip().lower() == 'no'):
                # if pd.isna(row[4]) :
                #     validated = False
                #     reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ', mobile number is empty.\n'
                #     #logfile.write(reason) 
                #     write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'mobile number is empty',getValue(row[1], str, ''))
                if not pd.isna(row[4]) and not bool(re.match(config.MOBILE_PATTERN, getMobileNumber(row[4],str,""))):
                    validated = False
                    reason = 'Water File data validation failed, Mobile number not correct for abas id '+ getValue(row[1], str, '') +'\n'
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Mobile number not correct',getValue(row[1], str, ''))
                    #logfile.write(reason)
                if pd.isna(row[5]):
                    validated = False
                    reason = 'Water File data validation failed for sl no. '+ getValue(row[0], str, '') + ',name is empty.\n'
                    #logfile.write(reason) 
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),' name is empty',getValue(row[1], str, ''))
                if not isna(row[8]) and getTime(row[8]) is None : 
                    validated = False
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),str(row[8])+'  Invalid DOB format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))

                # elif not pd.isna(row[5]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[5]))):
                #     validated = False
                #     reason = 'Sewerage File data validation failed, Name has invalid characters for sl no. '+ getValue(row[0], int, '') +'\n'
                #     #logfile.write(reason)  
                #     write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Name has invalid characters',getValue(row[1], str, ''))
                # if not pd.isna(row[9]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[9]))):                        
                #     validated = False
                #     reason = 'Water File data validation failed, Guardian Name has invalid characters for abas id '+ sgetValue(row[1], str, '') +'\n'
                #     write(logfile,waterFile,water_sheet.title,None,'Guardian Name has invalid characters',getValue(row[1], str, ''))
                #     #logfile.write(reason)

                if len(getValue(row[6], str, "")) > 0 and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[6]))) :                      
                    validated = False
                    reason = 'Water File data validation failed, Email id is not proper for abas id '+ getValue(row[1], str, '') +'\n'
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'Email id is not proper',getValue(row[1], str, ''))
                    #logfile.write(reason)
            if not pd.isna(row[1]):
                abasid = getValue(row[1], str, '')
                if str(abasid).strip() not in abas_ids:                    
                    validated = False
                    reason = 'there is no abas id available in property data for water connection sl no. '+ getValue(row[0], str, '') +'\n'
                    #logfile.write(reason) 
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'ABAS id not available in property data',getValue(row[1], str, ''))
                else:
                    if str(row[3]).strip().lower() == 'yes' and not pd.isna(abasid):
                        for obj in property_owner_obj[abasid]:
                            if(getValue(obj['ownerType'],str,"").lower() == 'multiple owners'):
                                validated = False
                                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),' Property ownership is multiple owner, so enter No for column D of water template and connection holder detail is mandatory ',getValue(row[1], str, ''))
            waterUsgType=getValue(row[24], str, '')
            if isna(waterUsgType):
                validated = False
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'water usage type is empty',getValue(row[1], str, ''))                   
            elif process_usage_type(waterUsgType,True) is None:
                validated = False
                write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'water usage type is not correct', getValue(row[1], str, ''))
            if waterUsgType.lower() == "residential" or  waterUsgType.lower() == "commercial" :
                waterSubusage = getValue(row[25], str, '')
                if waterSubusage != '' and waterSubusage.lower() not in  USAGE_SUB_USAGE_MAP[process_usage_type(waterUsgType,True)] :
                    validated = False            
                    # print("useage type ",subUsageValue.lower().strip())    
                    # print(USAGE_SUB_USAGE_MAP[process_usage_type(propUsgType,True)])        
                    write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'sub usage "'+waterSubusage+'" not correct as per usage type '+waterUsgType,getValue(row[1], str, ''))
            # if process_subusage_type(waterUsgType,True) is None:
            #     validated = False
            #     write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),'water sub usage type is not correct', getValue(row[1], str, ''))
        except Exception as ex:
            # write(logfile,waterFile,water_sheet.title,getValue(row[0], int, ''),str(ex) ,getValue(row[1], str, ''))
            print(config.CITY_NAME," validateWaterData Exception: ", getValue(row[0], int, ''), '  ', ex)
            traceback.print_exc()

    for index in range(3, water_sheet.max_row +1):
        try:
            if pd.isna(water_sheet['B{0}'.format(index)].value):                    
                break
            oldConnectionNo = getValue(water_sheet['C{0}'.format(index)].value, str,'')
            old_connections.append(str(oldConnectionNo).strip())
        except Exception as ex:
            print( config.CITY_NAME,  " validateDataForWater Exception: existing water connection no is empty: ",ex)
            traceback.print_exc()
            # write(logfile,waterFile,water_sheet.title,row[0],' existing water connection no is empty',getValue(row[1], str, ''))
    duplicate_ids = [item for item, count in collections.Counter(old_connections).items() if count > 1]

    if(len(duplicate_ids) >= 1):
        validated = False
        reason = 'Water File data validation failed. ' +'Duplicate existing water connection no for '+ str(duplicate_ids) +'\n'
        # print(reason)
        write(logfile,waterFile,water_sheet.title,None,'Duplicate existing water connection no for '+ str(duplicate_ids))
        #logfile.write(reason)   

    reason = 'Water file validation ends.\n'
    #print(reason)
    #logfile.write(reason) 
    return validated

def ValidateCols(waterFile, sheet, logfile):
    proper_column_order =['SI.No', 'Existing Property ID* ( Unique Value on which property are getting searched in existing system ) ',
     'Existing Water Consumer Number* ( Unique Value on which connection are getting searched in existing system ) ',
      "\n\nSelect 'Yes' if connection holder details are same as Property Owner and ownership type is single owner. In case of 'No' Fill the column E to N *",
       'Connection Holder Details', None, None, None, None, None, None, None, None, None, 'Pipe Size (inch)*', 'Water Source Type*', 'Connection Type*', 'Motor Info*', 'Activation Date*', 'Connection Permission*', 'Meter ID', 'Last Meter Reading ', 'Last Billed Date*', 'No.of taps*']
    
    column_list = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=2))]
    # print(len(proper_column_order))
    # print(len(column_list))
    validated = True
    try:
        ## Approach 1 : check for individual column 
        for i in range(0, 23):
            if(i== 3):
                continue
            if(proper_column_order[i] != column_list[i]) :
                print('Water file', column_list[i])
                validated = False
                write(logfile,waterFile,sheet.title,None,'Column order / name is not correct',column_list[i])
                # break
        

        ## Approach 2 -- Directly check difference in column
        missingColumnHeader = list(set(proper_column_order)-set(column_list))
        extraColumnHeader = list(set(column_list)-set(proper_column_order))
        if len (missingColumnHeader) > 0 : 
            write(logfile,waterFile,sheet.title,None,' Some Columns are Missing In Sheet')
            validated=False
        if len (extraColumnHeader) > 0 :
            write(logfile,waterFile,sheet.title,None,'Extra Columns exist in Sheet')
            validated=False
    except Exception as ex:
        validated = False
        print(config.CITY_NAME," validateCols Water Exception: ",ex)
        traceback.print_exc()
    return validated   

def createWaterJson(propertySheet, waterSheet, cityname, logfile, root, name):
    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0
    propertyNotAvailableCount = 0
    propertyNotAvailableArr = []
    auth_token = superuser_login()["access_token"]
    owner_obj = {}
    for i in range(3, propertySheet.max_row +1):  
        try:      
            abas_id = getValue(propertySheet['B{0}'.format(i)].value,str,None)
            if not pd.isna(abas_id):
                for row in propertySheet.iter_rows(min_row=i, max_col=42, max_row=i,values_only=True):                    
                    owner = Owner()
                    address = Address()
                    address.buildingName = getValue(row[17],str,"")
                    address.doorNo = getValue(row[18],str,"")
                    correspondence_address = get_propertyaddress(address.doorNo,address.buildingName,getValue(row[13],str,"Others"),cityname)
                    owner.name = getValue(row[28],str,"NAME")
                    owner.mobileNumber = getValue(row[29],str,"3000000000")
                    owner.emailId = getValue(row[30],str,"")
                    owner.gender = process_gender(row[31])
                    if not pd.isna(row[32]):
                        owner.dob = getTime(row[32])
                    owner.fatherOrHusbandName = getValue(row[33],str,"Guardian")
                    owner.relationship =  process_relationship(row[34])
                    if getValue( row[35] ,str,"Yes").lower() == "yes":
                        owner.sameAsPeropertyAddress = True   
                    else:
                        owner.sameAsPeropertyAddress = False 
                    if(owner.sameAsPeropertyAddress ==  True):
                        owner.correspondenceAddress = correspondence_address
                    else: 
                        owner.correspondenceAddress = getValue(row[36],str,correspondence_address)
                    owner.ownerType =  process_special_category(row[37])
                    if abas_id not in owner_obj:
                        owner_obj[abas_id] = []
                    owner_obj[abas_id].append(owner)
        except Exception as ex:
            traceback.print_exc()
            print(config.CITY_NAME," createWaterJson Exception: ", getValue(row[0], int, ''), '   ', ex)

    index = 2
    for row in waterSheet.iter_rows(min_row=3, max_col=26, max_row=waterSheet.max_row , values_only=True):
        
        index = index + 1
        abasPropertyId =  getValue(row[1],str,None)  
        
        property = Property() 
        
        tenantId = 'pb.'+ cityname
        property.tenantId = tenantId
        if pd.isna(abasPropertyId):
            print("empty Abas id in water file for sl no. ", row[0])
            break
        waterConnection = WaterConnection()
        connectionHolder = ConnectionHolder()
        processInstance = ProcessInstance()
        additionalDetail = AdditionalDetail()
        waterConnection.connectionHolders = []
        waterConnection.oldConnectionNo = getValue(row[2],str,None)

        status, res = waterConnection.search_water_connection(auth_token, tenantId, waterConnection.oldConnectionNo)               
        # with io.open(os.path.join(root, name,waterConnection.oldConnectionNo+"water_search_res.json"), mode="w", encoding="utf-8") as f:
        #     json.dump(res, f, indent=2,  ensure_ascii=False)  
        
        if(len(res['WaterConnection']) == 0):                
            status, res = property.search_abas_property(auth_token, tenantId, abasPropertyId)        
            # with io.open(os.path.join(root, name,"property_search_res.json"), mode="w", encoding="utf-8") as f:
            #     json.dump(res, f, indent=2,  ensure_ascii=False) 
            propertyId = ''
            if(len(res['Properties']) >= 1):
                for found_index, resProperty in enumerate(res["Properties"]):
                    propertyId = resProperty["propertyId"]
                    break
                
                try:  
                    print(cityname, "water",abasPropertyId)   
                    if(str(row[3]).strip().lower() == 'yes'):
                        waterConnection.connectionHolders = None
                        # owner = owner_obj[abasPropertyId]
                        # connectionHolder.name = owner.name
                        # connectionHolder.mobileNumber = owner.mobileNumber
                        # connectionHolder.fatherOrHusbandName = owner.fatherOrHusbandName
                        # connectionHolder.emailId = owner.emailId
                        # connectionHolder.correspondenceAddress = owner.correspondenceAddress
                        # connectionHolder.relationship = owner.relationship
                        # connectionHolder.ownerType = owner.ownerType
                        # connectionHolder.gender = owner.gender
                        # connectionHolder.sameAsPropertyAddress = True
                    else:
                        connectionHolder.name = getValue(row[5],str,"NAME")
                        connectionHolder.mobileNumber = getMobileNumber(row[4],str,"3000000000")
                        connectionHolder.fatherOrHusbandName = getValue(row[9],str,"Guardian")
                        connectionHolder.emailId = getValue(row[6],str,"")
                        connectionHolder.correspondenceAddress = getValue(row[12],str,"Correspondence")
                        connectionHolder.relationship = process_relationship(row[10])
                        connectionHolder.ownerType = process_special_category(row[13])
                        connectionHolder.gender = process_gender(row[7])
                        connectionHolder.sameAsPropertyAddress = False
                        waterConnection.connectionHolders.append(connectionHolder)
                    
                    
                    waterConnection.pipeSize = getValue(row[14],float,0.25)
                    waterConnection.proposedPipeSize = getValue(row[14],float,0.25)
                    waterConnection.waterSource = process_water_source(row[15])
                    if(waterConnection.waterSource != 'OTHERS'):
                        waterConnection.waterSubSource = waterConnection.waterSource.split('.')[1]                
                    else:
                        waterConnection.waterSubSource = ''
                        waterConnection.sourceInfo = 'Other'
                    waterConnection.connectionType = process_connection_type(row[16])
                    waterConnection.motorInfo  = process_motor_info(row[17])
                    waterConnection.propertyOwnership  = process_propertyOwnership(row[11])
                    waterConnection.authorizedConnection = process_connection_permission(row[19])
                    waterConnection.noOfTaps = getValue(row[23],int,1)
                    waterConnection.proposedTaps = getValue(row[23],int,1)
                    if( waterConnection.connectionType == 'Metered'):
                        waterConnection.meterId = getValue(row[20],str,None)
                        additionalDetail.initialMeterReading = getValue(row[21],float,None)
                    if not pd.isna(row[18]):
                        waterConnection.connectionExecutionDate = getTime(row[18])
                    additionalDetail.locality = ''
                    waterConnection.additionalDetails = additionalDetail
                    processInstance.action = 'ACTIVATE_CONNECTION'
                    waterConnection.tenantId = tenantId
                    waterConnection.propertyId = propertyId
                    waterConnection.processInstance = processInstance
                    waterConnection.water = True
                    waterConnection.sewerage = False
                    waterConnection.service = 'Water'
                    waterConnection.applicationType = 'NEW_WATER_CONNECTION'
                    waterConnection.applicationStatus = 'CONNECTION_ACTIVATED'
                    waterConnection.source = 'MUNICIPAL_RECORDS'
                    waterConnection.channel = 'DATA_ENTRY'
                    waterConnection.status = 'ACTIVE'
                    
                except Exception as ex:
                    traceback.print_exc()
                    print("createWaterJson Exception: ", getValue(row[0], int, ''), '   ', ex)

                req_data, statusCode, res = waterConnection.upload_water(auth_token, tenantId, waterConnection.oldConnectionNo, root, name)
                # with io.open(os.path.join(root, name,"water_create_res.json"), mode="w", encoding="utf-8") as f:
                #     json.dump(res, f, indent=2,  ensure_ascii=False)  
                if(statusCode == 200 or statusCode == 201):
                    for found_index, resWater in enumerate(res["WaterConnection"]):
                        connectionNo = resWater["connectionNo"]
                        # value = 'B{0}'.format(index) + '    ' + str(connectionNo) + '\n'
                        # logfile.write(value)
                        waterSheet['AA{0}'.format(index)].value = connectionNo
                        reason = 'water connection created for existing connection no. ' + str(waterConnection.oldConnectionNo)
                        # logfile.write(reason)
                        # print(reason)
                        createdCount = createdCount + 1
                        break
                else:
                    with io.open(os.path.join(root, name, abasPropertyId + "_water_create_req.json"), mode="w", encoding="utf-8") as f:
                        json.dump(req_data, f, indent=2,  ensure_ascii=False)
                    with io.open(os.path.join(root, name, abasPropertyId + "_water_create_res.json"), mode="w", encoding="utf-8") as f:
                        json.dump(res, f, indent=2,  ensure_ascii=False)
                    reason = 'water not created status code '+ str(statusCode)+ ' for existing connection no. ' + str(waterConnection.oldConnectionNo) + ' response: '+ str(res) + '\n'
                    # logfile.write(reason)
                    print(reason)
                    notCreatedCount = notCreatedCount + 1
            else:
                reason = 'property does not exist for abas id '+ abasPropertyId + '\n'
                propertyNotAvailableCount = propertyNotAvailableCount + 1
                propertyNotAvailableArr.append(abasPropertyId)
                print(reason)
                # logfile.write(reason)
        else:
            for found_index, resWater in enumerate(res["WaterConnection"]):
                connectionNo = resWater["connectionNo"]
                break
            waterSheet['AA{0}'.format(index)].value = connectionNo                    
            reason = 'water connection exist for existing connection no. ' + str(waterConnection.oldConnectionNo)
            # logfile.write(reason)
            # print(reason)
            searchedCount = searchedCount + 1
            

    reason = 'Water created count: '+ str(createdCount)
    print(reason)
    reason = 'Water not created count: '+ str(notCreatedCount)
    print(reason)
    reason = 'Water searched count: '+ str(searchedCount)
    print(reason)
    reason = 'Property not available count: '+ str(propertyNotAvailableCount)
    print(reason)
    print("Property not available arr: ", str(propertyNotAvailableArr))

def get_propertyaddress(doorNo, buildingName,locality,cityname):
    return doorNo + ' ' + buildingName + ' ' +locality + ' ' + cityname

def process_water_source(value):
    if value is None : 
        value ="Pipe-Treated"
    value = value.strip().lower()
    water_source_MAP = {
        'Ground-Borewell': 'GROUND.BOREWELL',
        'Ground-Handpump':'GROUND.HANDPUMP',
        'Ground-Well':'GROUND.WELL',
        'Surface-River':'SURFACE.RIVER',
        'Surface-Canal':'SURFACE.CANAL',
        'Surface-Lake':'SURFACE.LAKE',
        'Surface-Pond':'SURFACE.POND',
        'Surface-Rainwater':'SURFACE.RAINWATER',
        'Surface-Recycled Water':'SURFACE.RECYCLEDWATER',
        'Pipe-Treated':'PIPE.TREATED',
        'Pipe-Raw':'PIPE.RAW',
        'MES': 'MES'
    }
    for key in water_source_MAP :
        if key.lower() == value : 
            return water_source_MAP[key]    

    return water_source_MAP['Pipe-Treated']

def process_relationship(value):
    if value is None : 
        value ="parent"
    value = value.strip().lower()
    relationship_MAP = {
        "parent": "PARENT",
        "spouse": "SPOUSE",
        "gurdian": "GUARDIAN",
        "guardian": "GUARDIAN",
        "none": "PARENT",
        "na": "PARENT"
    }
    return relationship_MAP[value]

def process_gender(value):
    if value is None : 
        value ="Male"
    value =str(value).strip().lower()
    gender_MAP = {
        "male": "MALE",
        "female": "FEMALE",
        "transgender": "TRANSGENDER",
        "none": "MALE"       
    }
    return gender_MAP[value]

def process_connection_type(value):
    value = value.strip().lower()
    connection_MAP = {
        "metered": "Metered",
        "non-metered": "Non Metered",
        "nonmetered": "Non Metered",
        "none": "Non Metered"     
    }
    return connection_MAP[value]

def process_motor_info(value):
    value = value.strip()
    motor_info_MAP = {
        "None": "WITHOUTPUMP",
        "NA": "WITHOUTPUMP",
        "na": "WITHOUTPUMP",
        "With Pump": "WITHPUMP",
        "Without Pump": "WITHOUTPUMP"
    }
    return motor_info_MAP[value]

def process_propertyOwnership(value):
    value = value.strip()
    propertyOwnership_MAP = {
        "None": None,
        "HOR": "HOR",
        "Tenant": "TENANT"
    }
    return propertyOwnership_MAP[value]

def process_connection_permission(value):
    value = value.strip()
    connection_permission_MAP = {
        "Authorized": "AUTHORIZED",
        "Unauthorized": "UNAUTHORIZED",
        "None": "AUTHORIZED"
    }
    return connection_permission_MAP[value]
    
def process_special_category(value):
    if value is None: 
        value ="None"
    value =value.strip().lower()
    special_category_MAP = {
        "freedom fighter": "FREEDOMFIGHTER",
        "widow": "WIDOW",
        "handicapped": "HANDICAPPED",
        "below poverty line": "BPL",
        "defense personnel": "DEFENSE",
        "employee/staff of cb": "STAFF",
        "none of the above": "NONE",
        "none":"NONE",
        "na":"NONE"
    }
    return special_category_MAP[value]

## As its static data so need to load it again
USAGE_MAP = {
        "Residential" : "RESIDENTIAL" ,
        "Commercial" : "COMMERCIAL" ,
        "Institutional" : "INSTITUTIONAL",
        "Industrial" : "INDUSTRIAL" ,
        "NonResidential" : "NONRESIDENTIAL" ,
        "Non Residential" : "NONRESIDENTIAL" ,
        "Mixed" : "MIXED" 
    }
USAGE_MAP = { k.strip().lower():USAGE_MAP[k] for k in USAGE_MAP}
def process_usage_type(value, isValidation =False):
    value =value.strip().lower()
    if isValidation : 
        if value in USAGE_MAP : 
            return USAGE_MAP[value.strip().lower()] 
        return None
    return USAGE_MAP[value]

def process_usage_type(value, isValidation =False):
    value =value.strip().lower()
    if isValidation : 
        if value in USAGE_MAP : 
            return USAGE_MAP[value.strip().lower()] 
        return None
    if value is None: 
        value ="residential"    
    for key in USAGE_MAP :
        if key == value : 
            return USAGE_MAP[key]  
    return USAGE_MAP["residential"]  

SUB_USAGE_MAP = {
        "Residential" : "RESIDENTIAL.RESIDENTIAL" ,
        "Others" : "RESIDENTIAL.OTHERS" ,
        "Slum" : "RESIDENTIAL.SLUM" ,
        "Commercial" : "COMMERCIAL.COMMERCIAL" ,
        "Hotels" : "COMMERCIAL.HOTELS",
        "None":   ''   
    }
SUB_USAGE_MAP = { k.strip().lower():SUB_USAGE_MAP[k] for k in SUB_USAGE_MAP}

USAGE_SUB_USAGE_MAP =dict()
for ele in USAGE_MAP :
    if USAGE_MAP[ele] not in USAGE_SUB_USAGE_MAP  : 
        USAGE_SUB_USAGE_MAP[USAGE_MAP[ele]]=list()
for sub_usg_key in SUB_USAGE_MAP : 
    sub_usg_val =SUB_USAGE_MAP[sub_usg_key]
    for ele in USAGE_SUB_USAGE_MAP :
        if sub_usg_val.startswith(ele) :
            USAGE_SUB_USAGE_MAP[ele].append(sub_usg_key.strip().lower())

def process_subusage_type(value, isValidation =False):
    if isValidation : 
        if value in SUB_USAGE_MAP : 
            return SUB_USAGE_MAP[value.strip().lower()] 
        return None
    if value is None: 
        value ="none"    
    for key in SUB_USAGE_MAP :
        if key.lower() == value : 
            return SUB_USAGE_MAP[key]   
    return SUB_USAGE_MAP["none"]   

if __name__ == "__main__":
    main()

        


