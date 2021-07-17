from common import *
import numpy as np
from config import config, getValue, getTime,isna
import io
import os 
import sys
from common import superuser_login
from WaterConnection import *
from WaterLegacyDemand import *
import pandas as pd
import openpyxl
import collections
import re
import traceback

now = datetime.now()
date_time = now.strftime("%d-%m-%Y") 
# FOLDER_PATH  =r'D:\eGov\Data\WS\Azure Insertion'
# FOLDER_PATH  =r'D:\eGov\Data\WS\UAT Insertion'
# FOLDER_PATH  =r'C:\Users\Administrator\Downloads\WaterSewerageTemplates'
FOLDER_PATH  =r'D:\eGov\Data\WS\Legacy Demand'
cityToSkip = ['Bakloh','Bareilly','Dagshai','Dalhousie','Deolali','Ferozepur','Jabalpur','Jammu','Jutogh','Kanpur',
                'Kasauli','Khasyol','Meerut','Nainital','Pachmarhi','Ramgarh','Secunderabad','Subathu', 'roorkee']

# cityToInclude = ['Landour','Lansdowne','Lebong','Lucknow','Mathura','Mhow','Morar','Nasirabad','Pune','Ranikhet','Saugor',
#                 'Shahjahanpur','Shillong','Varanasi','Wellington']
# cityToInclude = ['Jalapahar','Jhansi','Kamptee','Lebong','Lucknow','Mathura']
cityToInclude = ['testing']


def main() :    
    print("Replace 109 of C:\ProgramData\Miniconda3\envs\py36\lib\site-packages\openpyxl\worksheet\merge.py with below one ") 
    print ("if side is None or  side.style is None:")
    # print('cityToSkip', len(cityToSkip))
    root = FOLDER_PATH    
    errorlogfile = open(os.path.join(root, "error CBs.txt"), "w")  
    successlogfile = open(os.path.join(root, "CB With ProperData.txt"), "w")
    notsuccesslogfile = open(os.path.join(root, "CB With ImProperData.txt"), "w")
    config.error_in_excel=[]
    config.error_in_multiple_owner=[]
    config.DATA_ENTRY_ISSUES_FOLDER =os.path.join(root,date_time + '-Data_Entries_Issues')
    config.DEMAMD_ENTRY_ISSUES_FOLDER =os.path.join(root,date_time + '-Demand_Entries_Issues')
    if not os.path.exists(config.DATA_ENTRY_ISSUES_FOLDER) :
        os.makedirs(config.DATA_ENTRY_ISSUES_FOLDER)
    if not os.path.exists(config.DEMAMD_ENTRY_ISSUES_FOLDER) :
        os.makedirs(config.DEMAMD_ENTRY_ISSUES_FOLDER)
    if not os.path.exists(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR")) :
        os.makedirs(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR"))
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        ####Only for some CBs
        # cityToInclude = getCitiesToInclude(cityToSkip,cb_module_data)
        for found_index, cityname in enumerate(cityToInclude):
            cityname =cityname.lower()
            config.errormsg=[]
            name = 'CB ' + cityname
            if  os.path.exists( os.path.join(root,name)):                
                try : 
                    if True:# cityname == 'jutogh' :
                        print("Processing for CB "+cityname.upper())
                        config.CITY_NAME = cityname
                        cbMain(cityname, successlogfile, notsuccesslogfile)
                except Exception as ex: 
                    print("Error in processing CB ",cityname , ex)
                    traceback.print_exc()
                    errorlogfile.write(cityname+"\n")
            if len(config.errormsg ) > 0 : 
                dateerror = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR",cityname+ "dateError.txt"), "w")  
                for element in config.errormsg:
                    dateerror.write(element + "\n") 
                dateerror.close()

        #### For all CBs
        # for found_index, module in enumerate(cb_module_data["tenants"]):
        #     if module["city"]["ulbGrade"]=="ST":
        #         continue
        #     cityname =module["code"].lower()[3:]
        #     config.errormsg=[]
        #     name = 'CB ' + cityname.lower()
        #     if  os.path.exists( os.path.join(root,name)):                
        #         try : 
        #             if True:# cityname == 'subathu' :
        #                 print("Processing for CB "+cityname.upper())
        #                 config.CITY_NAME = cityname
        #                 cbMain(cityname, successlogfile, notsuccesslogfile)
        #         except Exception as ex: 
        #             print("Error in processing CB ",cityname , ex)
        #             traceback.print_exc()
        #             errorlogfile.write(cityname+"\n")
        #     if len(config.errormsg ) > 0 : 
        #         dateerror = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"DATE_ERROR",cityname+ "dateError.txt"), "w")  
        #         for element in config.errormsg:
        #             dateerror.write(element + "\n") 
        #         dateerror.close()
        
    errorlogfile.close()
    successlogfile.close()   
    if len(config.error_in_excel) > 0 :   
        cbHaveExcelIssue = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"_CB_HAVE_EXCEL_ISSUE.txt"), "w")        
        for element in config.error_in_excel:
            cbHaveExcelIssue.write(element + "\n") 
        cbHaveExcelIssue.close()
     
    if len(config.error_in_multiple_owner) > 0 :     
        cbHaveMultipleOwnerIssue = open(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER,"_CB_HAVE_MULTIPLE_OWNER_ISSUE.txt"), "w")     
        for element in config.error_in_multiple_owner:
            cbHaveMultipleOwnerIssue.write(element + "\n") 
        cbHaveMultipleOwnerIssue.close()

def getCitiesToInclude(cityToSkip,cb_module_data):    
    cityToSkipLower = []
    for found_index, cityname in enumerate(cityToSkip):
        cityToSkipLower.append(cityname.lower())
    cityToSkipLower.sort()
    allCities = []
    cityToInclude = []
    for found_index, module in enumerate(cb_module_data["tenants"]):
        if module["city"]["ulbGrade"]=="ST":
            continue
        cityname =module["code"].lower()[3:]
        allCities.append(cityname)    
    try:
        allCities.sort()
        cityToInclude = np.setdiff1d(allCities, cityToSkipLower)
        # cityToInclude = set(allCities.sort()) - set(cityToSkipLower.sort())            
    except:
        traceback.print_exc() 
    return cityToInclude

def cbMain(cityname, successlogfile,notsuccesslogfile):
    Flag =False
    tenantMapping={}
    count = 0
    with io.open(config.TENANT_JSON, encoding="utf-8") as f:
        cb_module_data = json.load(f)
        for found_index, module in enumerate(cb_module_data["tenants"]):
            if module["city"]["ulbGrade"]=="ST":
                continue
            tenantMapping[module["code"].lower()]=module["code"].lower()[3:]

    # Doing for one cb at a time
    root = FOLDER_PATH
    name = 'CB ' + cityname.lower()
    demandFile =os.path.join(root, name,'Legacy Demand-' + cityname + '.xlsx')
    logfile = open(os.path.join(root, name, "Logfile.json"), "w")   
    connlogfile = open(os.path.join(root, name, "ConnLogfile.json"), "w")   
    countfile = open(os.path.join(root, name, "count.txt"), "w")  
    logfile.write("[ ")
    connlogfile.write("[ ")
    if os.path.exists(demandFile) : 
        validate =  validateDataForDemand(demandFile, logfile, cityname)
        if(validate == False):                
            print('Data validation for Demand Failed, Please check the log file.') 
            if config.INSERT_DATA: 
                return
        else:
            print('Data validation for Demand success.')                
        wb_demand = openpyxl.load_workbook(demandFile) 
        sheet1 = wb_demand.get_sheet_by_name('Demand')        
        
        if config.INSERT_DATA:
            createDemands(sheet1, cityname, logfile, root, name, countfile, connlogfile)                
            wb_demand.save(demandFile)        
        wb_demand.close()
    else:
        print("Deamand File doesnot exist for ", cityname) 

    logfile.seek(logfile.tell() - 1, os.SEEK_SET)
    logfile.write('')
    logfile.write("]")  

    connlogfile.seek(connlogfile.tell() - 1, os.SEEK_SET)
    connlogfile.write('')
    connlogfile.write("]")        

    size = os.path.getsize(os.path.join(root, name, "Logfile.json")) 
    logfile.close()
    size1 = os.path.getsize(os.path.join(root, name, "ConnLogfile.json")) 
    connlogfile.close() 
    try :  
        if size > 2 : 
            df = pd.read_json (os.path.join(root, name, "Logfile.json"))
            notsuccesslogfile.write(cityname)
            notsuccesslogfile.write("\n")  
            df.to_excel(os.path.join(config.DATA_ENTRY_ISSUES_FOLDER ,    name + " Data Entries Issues.xlsx"), index = None)
        elif size1 > 2 : 
            df = pd.read_json (os.path.join(root, name, "ConnLogfile.json"))
            notsuccesslogfile.write(cityname)
            notsuccesslogfile.write("\n")  
            df.to_excel(os.path.join(config.DEMAND_ENTRY_ISSUES_FOLDER ,    name + " Demand Entries Issues.xlsx"), index = None)
        else : 
            successlogfile.write(cityname)
            successlogfile.write("\n")            
    except Exception as ex: 
        print("Error in parsing json file",ex)

def validateDataForDemand(propertyFile, logfile, cityname):
    validated = True
    reason = ''
    try:
        wb_property = openpyxl.load_workbook(propertyFile) 
        sheet1 = wb_property.get_sheet_by_name('Demand')               
        validated = ValidateCols(logfile, propertyFile, sheet1)
        if not validated :
            print("Column Mismatch, sheets needs to be corrected")
            config["error_in_excel"].append(cityname +" have column issue in property sheet")

        # print('no. of rows in Property file sheet 1: ', sheet2.max_row ) 
        emptyRows=0
        count =0 
        for row in sheet1.iter_rows(min_row=2, max_col=5, max_row=sheet1.max_row ,values_only=True): 
            try : 
                if isna(row[1]) :
                    continue
                if isna(row[1]):
                    validated = False
                    reason = 'old connection number is empty'
                    writeDemandLog(logfile,getValue(row[0], int, ''),reason, getValue(row[1], str, ''))
                if not  isna(row[2]):
                    if  getValue(row[2], float, 0) < 0.0 or not bool(re.match("^\d*(\.[0-9]{1,2})*$",getValue(row[2], str, ''))):
                        validated = False
                        reason = 'water charge is not correct'
                        writeDemandLog(logfile,getValue(row[0], int, ''),reason, getValue(row[1], str, ''))
                if not  isna(row[3]):
                    if  getValue(row[3], float, 0) < 0.0 or not bool(re.match("^\d*(\.[0-9]{1,2})*$",getValue(row[3], str, ''))):
                        validated = False
                        reason = 'advance is not correct'
                        writeDemandLog(logfile,getValue(row[0], int, ''),reason, getValue(row[1], str, ''))
            except Exception as ex:
                print(config.CITY_NAME," validateDataForDemand Exception: ",getValue(row[0], str, ''), '  ',ex)
                traceback.print_exc()

        connection_ids = []
        for index in range(2, sheet1.max_row +1):
            try: 
                if isna(sheet1['B{0}'.format(index)].value):                    
                    break
                connectionId = getValue(sheet1['B{0}'.format(index)].value, str, '')
                connection_ids.append(connectionId)
            except Exception as ex:
                traceback.print_exc()
                print( config.CITY_NAME,  " validateDataForProperty Exception: abas id is empty: ",getValue(row[0], int, ''), '  ',ex)
        duplicate_ids = [item for item, count in collections.Counter(connection_ids).items() if count > 1]

        # if(len(duplicate_ids) >= 1):
        #     validated = False
        #     write(logfile,None,'Duplicate old connection id for '+ str(duplicate_ids))  
    except Exception as ex:
        traceback.print_exc()
    return validated

def ValidateCols(logfile, propertyFile, sheet1):
    proper_column_order1 = ['Sl No', 'Old Connection No','Water charge','Penalty','Advance']
    validated = True
    column_list = [c.value for c in next(sheet1.iter_rows(min_row=1, max_row=1))]
    try:
        for i in range(0, 4):
            if(proper_column_order1[i].strip() != column_list[i].strip()) :
                print('Demand file: ', column_list[i])
                validated = False
                write(logfile,propertyFile,sheet1.title,None,'Column order / name is not correct',column_list[i])
                # break

    except Exception as ex:
        validated = False
        print(config.CITY_NAME," validateCols Demand Exception: ",ex)
        traceback.print_exc()
    return validated  


def createDemands(sheet1, cityname, logfile,root, name, countfile, connlogfile):

    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0
    auth_token = superuser_login()["access_token"]
    tenantId = 'pb.'+ cityname
    
    index = 2
    waterDemandsObj = []  
    demands = WaterDemands()  
    for row in sheet1.iter_rows(min_row=2, max_col=6, max_row=sheet1.max_row ,values_only=True): 
        waterDemand = WaterDemand()  
        waterConnection = WaterConnection()
        index = index + 1  
        try:
            if isna(getValue(row[1], str, None)) :
                continue
            waterDemand.oldConnectionNo= getValue(row[1], str, '')
            status, res = waterConnection.search_water_connection(auth_token, tenantId, waterDemand.oldConnectionNo)                        
            if(len(res['WaterConnection']) > 0): 
                # print(getValue(row[2], float, 0))
                if not (isna(row[2]) and isna(row[3]) and isna(row[4])) and not (getValue(row[2], int, 0) == 0 and getValue(row[3], int, 0) == 0 and getValue(row[4], int, 0) == 0) :   
                    waterDemand.waterCharge = getValue(row[2], float, 0)
                    waterDemand.penalty = getValue(row[3], float, 0)            
                    waterDemand.advance = getValue(row[4], float, 0)            
                    waterDemandsObj.append(waterDemand)
        except Exception as ex:
            print(config.CITY_NAME," createDemands Exception: ",ex)
            traceback.print_exc() 
    demands.waterDemands = waterDemandsObj
    statusCode, res = demands.upload_demand(auth_token, tenantId, demands, root, name)
    if len(res) > 0 :
        errorDemandConn = open(os.path.join(config.DEMAMD_ENTRY_ISSUES_FOLDER, name + "Demand Entry Issues.xlsx"), "w")  
        for connection in res:
            writeErrorDemandConn(connlogfile, getValue(connection, str, ''))
    

    # with io.open(os.path.join(root, name,str(property.abasPropertyId) + "_property_create_res.json"), mode="w", encoding="utf-8") as f:
    #     json.dump(res, f, indent=2,  ensure_ascii=False)
    # if statusCode == 200 :
    #     for found_index, resProperty in enumerate(res["Properties"]):
    #         propertyId = resProperty["propertyId"]                    
    #         createdCount = createdCount + 1
    #         break
    # elif statusCode == 400 :
    #     with io.open(os.path.join(root, name,  "failedConnections.json"), mode="w", encoding="utf-8") as f:
    #         json.dump(res, f, indent=2,  ensure_ascii=False)                
    #     reason = 'Demand not created status code '+ str(statusCode) + ' for connection id ' + str(property.abasPropertyId) + ' response: ', str(res)  + '\n'
    #     print(reason)
    #     notCreatedCount = notCreatedCount + 1
     
    reason = 'Demand created count: '+ str(createdCount)
    print(reason)
    countfile.write(reason)
    countfile.write('\n')
    reason = 'Demand not created count: '+ str(notCreatedCount)
    print(reason)
    countfile.write(reason)
    countfile.write('\n')
    reason = 'Demand searched count: '+ str(searchedCount)
    print(reason)
    countfile.write(reason)
    countfile.write('\n')

if __name__ == "__main__":
    # print(USAGE_MAP)
    main()
    
    
              


