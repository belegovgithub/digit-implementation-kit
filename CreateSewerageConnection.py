from common import *
from config import config, getValue, getMobileNumber, getTime, isna
import io
import os 
import sys
from SewerageConnection import *
from PropertyTax import *
import pandas as pd
# from CreateProperty import getValue
import openpyxl
import collections
import traceback
    
def ProcessSewerageConnection(propertyFile, sewerageFile, logfile, root, name,  cityname, property_owner_obj = {}) :
    wb_property = openpyxl.load_workbook(propertyFile) 
    propertySheet = wb_property.get_sheet_by_name('Property Assembly Detail') 
    wb_sewerage = openpyxl.load_workbook(sewerageFile) 
    sewerageSheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details')  
    #print('no. of rows in sewerage file: ', sewerageSheet.max_row +1 ) 
    validate = validateSewerageData(propertySheet, sewerageFile, logfile, cityname, property_owner_obj)  
    if(validate == False):                
        print('Data validation for sewerage Failed, Please check the log file.') 
        if config.INSERT_DATA: 
            return
    else:
        print('Data validation for sewerage success.')
    if config.INSERT_DATA and config.CREATE_SEWERAGE: 
        createSewerageJson(propertySheet, sewerageSheet, cityname, logfile, root, name)   
        wb_sewerage.save(sewerageFile)        
    wb_sewerage.close()

def validateSewerageData(propertySheet, sewerageFile, logfile, cityname, property_owner_obj):
    validated = True
    wb_sewerage = openpyxl.load_workbook(sewerageFile) 
    sewerage_sheet = wb_sewerage.get_sheet_by_name('Sewerage Connection Details') 
    index = 2
    reason = 'sewerage file validation starts.\n'
    # validated = ValidateCols(sewerageFile, sewerage_sheet, logfile)
    # if not validated :
    #     print("Column Mismatch, sheets needs to be corrected")
    #     config["error_in_excel"].append(cityname +" have issue in Sewerage sheet")

    abas_ids = [] 
    old_connections = []
    try:
        for index in range(3, propertySheet.max_row +1 ):
            if pd.isna(propertySheet['B{0}'.format(index)].value):
                # validated = False
                # reason = 'Sewerage File data validation failed, Sl no. column is empty'
                # #logfile.write(reason)
                # write(logfile,"property excel",propertySheet.title,index,'Sl no. column is empty')
                continue
            propSheetABASId = getValue(propertySheet['B{0}'.format(index)].value, str, '') 
            abas_ids.append(str(propSheetABASId).strip())   
    except Exception as ex:
        print(config.CITY_NAME," validateSewerageData Exception: ", ex)
    emptyRows =0 
    for row in sewerage_sheet.iter_rows(min_row=3, max_col=22, max_row=sewerage_sheet.max_row +1 ,values_only=True):        
        index = index + 1
        try:
            # if emptyRows > 10 :
            #     break
            if pd.isna(row[1]) and pd.isna(row[2]):
                emptyRows = emptyRows +1
                continue
            if pd.isna(row[0]):
                validated = False
                reason = 'Sewerage File data validation failed, Sl no. column is empty\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'Sl no. column is empty',getValue(row[1], str, ''))
            if pd.isna(row[1]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ getValue(row[0], str, '') + ', abas id is empty.\n'
                #logfile.write(reason) 
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'abas id is empty',getValue(row[1], str, ''))
            if pd.isna(row[2]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ getValue(row[0], str, '') + ', existing sewerage connection number is empty.\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'existing sewerage connection number is empty',getValue(row[1], str, ''))

            if pd.isna(row[3]):
                validated = False
                reason = 'Sewerage File data validation failed for sl no. '+ getValue(row[0], str, '') + ', same as property address cell is empty.\n'
                #logfile.write(reason)
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'same as property address cell is empty',getValue(row[1], str, ''))
            if(str(row[3]).strip().lower() == 'no'):
                # if pd.isna(row[4]) :
                #     validated = False
                #     reason = 'Sewerage File data validation failed for sl no. '+ getValue(row[0], str, '') + ', mobile number is empty.\n'
                #     #logfile.write(reason) 
                #     write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'mobile number is empty',getValue(row[1], str, ''))
                if not pd.isna(row[4]) and ( len(getMobileNumber(row[4],str,"")) != 10):
                        validated = False
                        reason = 'Sewerage File data validation failed, Mobile number not correct for abas id '+ str(getValue(row[0], int, '')) +'\n'
                        write(logfile,sewerageFile,sewerage_sheet.title,None,'Mobile number not correct',getValue(row[0], int, ''))
                        #logfile.write(reason)
                if pd.isna(row[5]):
                    validated = False
                    reason = 'Sewerage File data validation failed for sl no. '+ str(getValue(row[0], int, '')) + ',name is empty.\n'
                    #logfile.write(reason) 
                    write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),' name is empty',getValue(row[1], str, ''))
                if not isna(row[8]) and getTime(row[8]) is None : 
                    validated = False
                    write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),str(row[8])+'  Invalid DOB format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))

                # elif not pd.isna(row[5]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[5]))):
                #     validated = False
                #     write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'Name has invalid characters',getValue(row[1], str, ''))
                # if not pd.isna(row[9]) and not bool(re.match("[a-zA-Z \\.]+$",str(row[9]))):                        
                #     validated = False
                #     write(logfile,sewerageFile,sewerage_sheet.title,None,'Guardian Name has invalid characters',getValue(row[0], int, ''))
                if len(getValue(row[6], str, "")) > 0 and not bool(re.match("^[a-zA-Z0-9+_.-]+@[a-zA-Z0-9]+.(([a-zA-Z\-0-9]+\.)+[a-zA-Z]{2,})$",str(row[6]))) :                      
                    validated = False
                    write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'Email id is not proper',getValue(row[1], str, ''))
            # if pd.isna(row[18]):
            #     validated = False
            #     write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'last billed date is empty',getValue(row[1], str, ''))
            # elif pd.isna(getTime(row[18])):
            #     validated = False
            #     write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),str(row[18]) +' Invalid Billing date format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))
            if pd.isna(row[17]):
                validated = False
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'Activation date is empty',getValue(row[1], str, ''))
            elif pd.isna(getTime(row[17])):
                validated = False
                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),str(row[17]) +' Invalid Activation date format,Valid format is : dd/mm/yyyy(24/04/2021) ',getValue(row[1], str, ''))    

            if not pd.isna(row[1]):
                abasid = getValue(row[1], str, '')
                if str(abasid).strip() not in abas_ids:
                    validated = False
                    reason = 'there is no abas id available in property data for sewerage connection sl no. '+ getValue(row[0], str, '') +'\n'
                    #logfile.write(reason) 
                    write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'ABAS id not available in property data',getValue(row[1], str, ''))
                else:
                    if str(row[3]).strip().lower() == 'yes' and not pd.isna(abasid):
                        for obj in property_owner_obj[abasid]:
                            if(getValue(obj['ownerType'],str,"") == 'Multiple Owners'):
                                validated = False
                                write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),' Property ownership is multiple owner, so enter No for column D of sewerage template and connection holder detail is mandatory ',getValue(row[1], str, ''))
            
        except Exception as ex:
            print(config.CITY_NAME," validateSewerageData Exception: ", getValue(row[0], int, ''), '   ', ex)
            # write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),str(ex) ,getValue(row[1], str, ''))
    for index in range(3, sewerage_sheet.max_row +1):
        try:
            if pd.isna(sewerage_sheet['B{0}'.format(index)].value):                    
                break
            oldConnectionNo = getValue(sewerage_sheet['C{0}'.format(index)].value, str,'')
            old_connections.append(str(oldConnectionNo).strip())
        except Exception as ex:
            print( config.CITY_NAME,  " validateDataForSewerage Exception: existing sewerage connection no is empty: ",ex)
            traceback.print_exc()
            write(logfile,sewerageFile,sewerage_sheet.title,getValue(row[0], int, ''),'existing sewerage connection no is empty',getValue(row[1], str, ''))
    duplicate_ids = [item for item, count in collections.Counter(old_connections).items() if count > 1]

    if(len(duplicate_ids) >= 1):
        validated = False
        reason = 'Sewerage File data validation failed. ' +'Duplicate existing sewerage connection no for '+ str(duplicate_ids) +'\n'
        # print(reason)
        write(logfile,sewerageFile,sewerage_sheet.title,None,'Duplicate existing sewerage connection no for '+ str(duplicate_ids))
        #logfile.write(reason)  
    reason = 'sewerage file validation ends.\n'
    #print(reason)
    #logfile.write(reason) 
    return validated

def ValidateCols(sewerageFile, sheet, logfile):
    proper_column_order = [None, None, None, "\n\nSelect 'Yes' if connection holder details are same as Property Owner and ownership type is single owner. In case of 'No' Fill the column E to N *", 
    'Mobile No*', 'Connection Holder Name*', 'Email Id', 'Gender *', 'DOB ', 'Guardian Name *',
     'Relationship *', 'Ownership Type *', 'Address *', 'Special Category *', 'Drainage Pipe Size (inch) *', 'No of Water Closets *', 'No of Toilets *', 'Activation Date *', 'Last Billed Date *']
    column_list = [c.value for c in next(sheet.iter_rows(min_row=2, max_row=2))]

    validated = True
    try:
        ## Approach 1 : check for individual column 
        for i in range(4, 19):
            if(proper_column_order[i].strip() != column_list[i].strip()) :
                print('Sewerage file', column_list[i])
                validated = False
                write(logfile,sewerageFile,sheet.title,None,'Column order / name is not correct',column_list[i])
                # break
        # Approach 2 -- Directly check difference in column
        
        missingColumnHeader = list(set(proper_column_order)-set(column_list))
        extraColumnHeader = list(set(column_list)-set(proper_column_order))
        if len (missingColumnHeader) > 0 : 
            write(logfile,sewerageFile,sheet.title,None,' Some Columns are Missing In Sheet')
            validated=False
        if len (extraColumnHeader) > 0 :
            write(logfile,sewerageFile,sheet.title,None,'Extra Columns exist in Sheet')
            validated=False
    except Exception as ex:
        validated = False
        print(config.CITY_NAME," validateCols Sewerage Exception: ",ex)
        traceback.print_exc()
    return validated 
    
def createSewerageJson(propertySheet, sewerageSheet, cityname, logfile, root, name):
    createdCount = 0
    searchedCount = 0
    notCreatedCount = 0
    auth_token = superuser_login()["access_token"]
    owner_obj = {}
    for i in range(3, propertySheet.max_row +1 ):    
        try:    
            abas_id = getValue(propertySheet['B{0}'.format(i)].value,str,None)
            if not pd.isna(abas_id):
                for row in propertySheet.iter_rows(min_row=i, max_col=42, max_row=i,values_only=True):                    
                    owner = Owner()
                    address = Address()
                    address.buildingName = getValue(row[17].strip(),str,"")
                    address.doorNo = getValue(row[18],str,"")
                    correspondence_address = get_propertyaddress(address.doorNo,address.buildingName,getValue(str(row[13]).strip(),str,"Others"),cityname)
                    owner.name = getValue(row[28],str,"NAME")
                    owner.mobileNumber = getValue(row[29],str,"3000000000")
                    owner.emailId = getValue(row[30],str,"")
                    owner.gender = process_gender(row[31])
                    owner.fatherOrHusbandName = getValue(row[33],str,"Guardian")
                    owner.relationship =  process_relationship(row[34])
                    owner.sameAsPeropertyAddress = getValue(str(row[35]).strip(),bool,True)
                    if(owner.sameAsPeropertyAddress ==  True):
                        owner.correspondenceAddress = correspondence_address
                    else: 
                        owner.correspondenceAddress = getValue(row[36],str,correspondence_address)
                    owner.ownerType =  process_special_category(str(row[37]).strip())
                    if abas_id not in owner_obj:
                        owner_obj[abas_id] = []
                    owner_obj[abas_id].append(owner)
        except Exception as ex:
            print(config.CITY_NAME," createSewerageJson Exception: ", getValue(row[0], int, ''), '   ', ex)
    index = 2
    for row in sewerageSheet.iter_rows(min_row=3, max_col=22, max_row=sewerageSheet.max_row +1 , values_only=True):
         
        index = index + 1
        abasPropertyId =  getValue(row[1],str,None)          
        property = Property() 
        
        tenantId = 'pb.'+ cityname
        property.tenantId = tenantId
        if pd.isna(abasPropertyId):
            print("empty Abas id in sewerage file for sl no. ", getValue(row[0], int, ''))
            break
        sewerageConnection = SewerageConnection()
        connectionHolder = ConnectionHolder()
        processInstance = ProcessInstance()
        additionalDetail = AdditionalDetail()
        sewerageConnection.connectionHolders = []
        sewerageConnection.oldConnectionNo = getValue(row[2],str,None)
        status, res = sewerageConnection.search_sewerage_connection(auth_token, tenantId, sewerageConnection.oldConnectionNo)               
        # with io.open(os.path.join(root, name,"sewerage_search_res.json"), mode="w", encoding="utf-8") as f:
        #     json.dump(res, f, indent=2,  ensure_ascii=False)  
        if(len(res['SewerageConnections']) == 0):
            
            status, res = property.search_abas_property(auth_token, tenantId, abasPropertyId)        
            # with io.open(os.path.join(root, name,"property_search_res.json"), mode="w", encoding="utf-8") as f:
            #     json.dump(res, f, indent=2,  ensure_ascii=False) 
            propertyId = ''
            if(len(res['Properties']) >= 1):
                for found_index, resProperty in enumerate(res["Properties"]):
                    propertyId = resProperty["propertyId"]
                    break
                try: 
                    print(cityname, "Sewerage",abasPropertyId)
                    if(str(row[3]).strip().lower() == 'yes'):
                        sewerageConnection.connectionHolders = None
                    else:
                        connectionHolder.name = getValue(row[5],str,"NAME")
                        connectionHolder.mobileNumber = getValue(row[4],str,"3000000000")
                        connectionHolder.fatherOrHusbandName = getValue(row[9],str,"Guardian")
                        connectionHolder.emailId = getValue(row[6],str,"")
                        connectionHolder.correspondenceAddress = getValue(row[12],str,"Correspondence")
                        connectionHolder.relationship = process_relationship(row[10])
                        connectionHolder.ownerType = process_special_category(str(row[13]).strip())
                        connectionHolder.gender = process_gender(row[7])
                        connectionHolder.sameAsPropertyAddress = False
                        sewerageConnection.connectionHolders.append(connectionHolder)
                    
                    
                    sewerageConnection.drainageSize = getValue(row[14],float,0.25)
                    sewerageConnection.proposedDrainageSize = getValue(row[14],float,0.25)                    
                    sewerageConnection.propertyOwnership  = process_propertyOwnership(str(row[11]))
                    sewerageConnection.noOfWaterClosets = getValue(row[15],int,1)
                    sewerageConnection.proposedWaterClosets = getValue(row[15],int,1)
                    sewerageConnection.noOfToilets = getValue(row[16],int,1)
                    sewerageConnection.proposedToilets = getValue(row[16],int,1)
                    if not pd.isna(row[17]):
                        sewerageConnection.connectionExecutionDate = getTime(row[17])
                    additionalDetail.locality = ''
                    sewerageConnection.additionalDetails = additionalDetail
                    processInstance.action = 'ACTIVATE_CONNECTION'
                    sewerageConnection.tenantId = tenantId
                    sewerageConnection.propertyId = propertyId
                    sewerageConnection.processInstance = processInstance
                    sewerageConnection.water = False
                    sewerageConnection.sewerage = True
                    sewerageConnection.service = 'Sewerage'
                    sewerageConnection.applicationType = 'NEW_SEWERAGE_CONNECTION'
                    sewerageConnection.applicationStatus = 'CONNECTION_ACTIVATED'
                    sewerageConnection.source = 'MUNICIPAL_RECORDS'
                    sewerageConnection.channel = 'DATA_ENTRY'
                    sewerageConnection.status = 'ACTIVE'
                except Exception as ex:
                    print("createSewerageJson Exception: ", getValue(row[0], int, ''), '   ', ex)

                
                req_data, statusCode, res = sewerageConnection.upload_sewerage(auth_token, tenantId, sewerageConnection.oldConnectionNo, root, name)
                # with io.open(os.path.join(root, name,"sewerage_create_res.json"), mode="w", encoding="utf-8") as f:
                #     json.dump(res, f, indent=2,  ensure_ascii=False)  
                sewerageconnectionNo = '' 
                if(statusCode == 200 or statusCode == 201):
                    for found_index, resSewerage in enumerate(res["SewerageConnections"]):
                        connectionNo = resSewerage["connectionNo"]
                        # value = 'B{0}'.format(index) + '    ' + str(connectionNo) + '\n'
                        # logfile.write(value)
                        sewerageSheet['T{0}'.format(index)].value = connectionNo
                        reason = 'sewerage connection created for existing connection no. ' + str(sewerageConnection.oldConnectionNo)
                        # logfile.write(reason)
                        # print(reason)
                        createdCount = createdCount + 1
                        break
                else:
                    with io.open(os.path.join(root, name, abasPropertyId + "_sewerage_create_req.json"), mode="w", encoding="utf-8") as f:
                        json.dump(req_data, f, indent=2,  ensure_ascii=False)
                    with io.open(os.path.join(root, name, abasPropertyId + "_sewerage_create_res.json"), mode="w", encoding="utf-8") as f:
                        json.dump(res, f, indent=2,  ensure_ascii=False)
                    reason = 'sewerage not created status code '+ str(statusCode)+ ' for existing connection no. ' + str(sewerageConnection.oldConnectionNo) + ' response: '+ str(res) + '\n'
                    # logfile.write(reason)
                    print(reason)
                    notCreatedCount = notCreatedCount + 1                      
            else:
                reason = 'property does not exist for abas id '+ str(property.abasPropertyId) + '\n'
                print(reason)
                # logfile.write(reason)
        else:
            for found_index, resSewerage in enumerate(res["SewerageConnections"]):
                connectionNo = resSewerage["connectionNo"]
                break
            sewerageSheet['T{0}'.format(index)].value = connectionNo
            reason = 'sewerage connection exist for existing connection no. ' + str(sewerageConnection.oldConnectionNo)
            # logfile.write(reason)
            # print(reason)
            searchedCount = searchedCount + 1
            

    reason = 'sewerage created count: '+ str(createdCount)
    print(reason)
    reason = 'sewerage not created count: '+ str(notCreatedCount)
    print(reason)
    reason = 'sewerage searched count: '+ str(searchedCount)
    print(reason)


def get_propertyaddress(doorNo, buildingName,locality,cityname):
    return doorNo + ' ' + buildingName + ' ' +locality + ' ' + cityname


def process_relationship(value):
    if value is None : 
        value ="parent"
    value = value.strip().lower()
    relationship_MAP = {
        "parent": "PARENT",
        "spouse": "SPOUSE",
        "gurdian": "GUARDIAN",
        "guardian": "GUARDIAN",
        "none": "PARENT",
        "na": "PARENT"
    }
    return relationship_MAP[value]

def process_gender(value):
    if value is None : 
        value ="Male"
    value =str(value).strip().lower()
    gender_MAP = {
        "male": "MALE",
        "female": "FEMALE",
        "transgender": "TRANSGENDER",
        "none": "MALE"       
    }
    return gender_MAP[value]

def process_connection_type(value):
    connection_MAP = {
        "Metered": "Metered",
        "Non-Metered": "Non Metered",
        "None": "Non Metered"     
    }
    return connection_MAP[value]

def process_propertyOwnership(value):
    propertyOwnership_MAP = {
        "None": None,
        "HOR": "HOR",
        "Tenant": "TENANT"
    }
    return propertyOwnership_MAP[value]

def process_connection_permission(value):
    connection_permission_MAP = {
        "Authorized": "AUTHORIZED",
        "Unauthorized": "UNAUTHORIZED",
        "None": "AUTHORIZED"
    }
    return connection_permission_MAP[value]

def process_special_category(value):
    special_category_MAP = {
        "Freedom fighter": "FREEDOMFIGHTER",
        "Widow": "WIDOW",
        "Handicapped": "HANDICAPPED",
        "Below Poverty Line": "BPL",
        "Defense Personnel": "DEFENSE",
        "Employee/Staff of CB": "STAFF",
        "None of the above": "NONE",
        "None":"NONE",
        "na":"NONE"
    }
    return special_category_MAP[value]


if __name__ == "__main__":
    main()    
    type         


